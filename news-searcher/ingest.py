#!/usr/bin/env python3
"""
ingest.py — Manual URL ingestion tool for the HIPAA Breach Intelligence database.

Fetches story URLs, uses a local Ollama model to extract structured story data,
and appends results to the Excel file used by searcher.py.

Self-contained — does not require importing from searcher.py.

Dependencies: pip install requests beautifulsoup4 openpyxl
Requires:     Ollama running locally (default: http://localhost:11434)

Usage examples:
  python ingest.py --urls https://example.com/story1 https://example.com/story2
  python ingest.py --file urls.txt
  python ingest.py --source https://www.techtarget.com/healthtechsecurity
  python ingest.py --urls urls.txt                        # auto-detected as a file
  python ingest.py --urls https://example.com --dry-run
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Config ────────────────────────────────────────────────────────────────────

EXCEL_FILE = "hipaa_breach_intelligence.xlsx"
SHEET_NAME = "Stories"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; ComplianceIntelligenceBot/1.0; research purposes)",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ── Excel schema (must match searcher.py) ─────────────────────────────────────

COLUMNS = [
    ("Date",               16, "date"),
    ("Category",           13, "category"),
    ("Relevance",          12, "small_biz_relevance"),
    ("Entity Size",        13, "entity_size"),
    ("Patients Affected",  18, "affected_count"),
    ("Title",              52, "title"),
    ("Summary",            72, "summary"),
    ("Source",             24, "source"),
    ("Penalty / Fine",     16, "penalty"),
    ("Tags",               32, "tags"),
    ("URL",                52, "url"),
    ("Run Date",           18, "run_date"),
    ("Model",              20, "model"),
]

_TITLE_COL = next(i + 1 for i, (_, _, k) in enumerate(COLUMNS) if k == "title")

HEADER_BG  = "1F3864"
HEADER_FG  = "FFFFFF"
HIGH_BG    = "FFF2CC"
HIGH_ALT   = "FDE9A2"
MID_BG     = "E2EFDA"
MID_ALT    = "D0E8C5"
LOW_BG     = "EDF2F9"
LOW_ALT    = "DAE4F2"
BORDER_COL = "BFBFBF"


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _thin_border() -> Border:
    s = Side(style="thin", color=BORDER_COL)
    return Border(left=s, right=s, top=s, bottom=s)


def _write_header(ws) -> None:
    ws.row_dimensions[1].height = 22
    for col_idx, (label, width, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"


def _story_to_values(story: dict, run_date: str, model: str) -> list:
    tags = story.get("tags") or []
    return [
        story.get("date") or "",
        (story.get("category") or "").title(),
        (story.get("small_biz_relevance") or "").title(),
        (story.get("entity_size") or "").title(),
        story.get("affected_count") or "",
        story.get("title") or "",
        story.get("summary") or "",
        story.get("source") or "",
        story.get("penalty") or "",
        " · ".join(tags) if isinstance(tags, list) else str(tags),
        story.get("url") or "",
        run_date,
        model,
    ]


def _row_bg(story: dict, row_num: int) -> str:
    relevance = story.get("small_biz_relevance", "low")
    even      = (row_num % 2 == 0)
    if relevance == "high":
        return HIGH_ALT if even else HIGH_BG
    if relevance == "medium":
        return MID_ALT if even else MID_BG
    return LOW_ALT if even else LOW_BG


def _write_data_row(ws, row_num: int, values: list, story: dict) -> None:
    bg        = _row_bg(story, row_num)
    url_col   = next(i + 1 for i, (_, _, k) in enumerate(COLUMNS) if k == "url")
    wrap_cols = {i + 1 for i, (_, _, k) in enumerate(COLUMNS) if k in ("title", "summary", "tags")}

    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.border    = _thin_border()
        cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in wrap_cols))
        if col_idx == url_col and value:
            cell.hyperlink = value
            cell.font = Font(name="Arial", size=9, color="1155CC", underline="single")
        else:
            cell.font = Font(name="Arial", size=9)
    ws.row_dimensions[row_num].height = 52


def save_or_append_excel(stories: list[dict], model: str, path: str) -> int:
    """Create or append to the Excel file. Returns number of rows written."""
    run_date  = datetime.now().strftime("%Y-%m-%d %H:%M")
    title_idx = _TITLE_COL - 1  # 0-based

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

        existing: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[title_idx] if len(row) > title_idx else None
            if val:
                existing.add(str(val).strip().lower())

        new_stories = [
            s for s in stories
            if str(s.get("title", "")).strip().lower() not in existing
        ]

        if not new_stories:
            print("  → No new stories (all duplicates of existing rows).")
            return 0

        ws.insert_rows(2, amount=len(new_stories))
        for i, story in enumerate(new_stories):
            row_num = 2 + i
            _write_data_row(ws, row_num, _story_to_values(story, run_date, model), story)

        wb.save(path)
        print(f"  → Appended {len(new_stories)} new stories to: {path}")
        return len(new_stories)

    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.sheet_properties.tabColor = "1F3864"
        _write_header(ws)

        for i, story in enumerate(stories):
            row_num = 2 + i
            _write_data_row(ws, row_num, _story_to_values(story, run_date, model), story)

        wb.save(path)
        print(f"  → Created: {path}  ({len(stories)} stories)")
        return len(stories)


# ── Fetching ──────────────────────────────────────────────────────────────────

def fetch_page(url: str, timeout: int = 20) -> str:
    """Fetch URL, strip chrome, return up to 8000 chars of visible text."""
    import urllib3

    def _extract(r) -> str:
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "header",
                          "aside", "form", "noscript", "iframe"]):
            tag.decompose()
        text = soup.get_text(separator="\n", strip=True)
        return re.sub(r"\n{3,}", "\n\n", text)[:8000]

    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        return _extract(r)
    except requests.exceptions.SSLError:
        try:
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
            r = requests.get(url, headers=HEADERS, timeout=timeout, verify=False)
            r.raise_for_status()
            return _extract(r)
        except Exception as e:
            return f"[SSL error: {e}]"
    except requests.exceptions.Timeout:
        return f"[Timeout: {url}]"
    except requests.exceptions.ConnectionError as e:
        return f"[Connection error: {e}]"
    except Exception as e:
        return f"[Error: {e}]"


def fetch_links_from_source(source_url: str, max_links: int = 20) -> list[str]:
    """Crawl an index page and return article URLs on the same domain."""
    print(f"  Scanning source: {source_url}", flush=True)
    try:
        r = requests.get(source_url, headers=HEADERS, timeout=20)
        r.raise_for_status()
    except Exception as e:
        print(f"  ⚠ Could not fetch source page: {e}")
        return []

    soup  = BeautifulSoup(r.text, "html.parser")
    base  = f"{urlparse(source_url).scheme}://{urlparse(source_url).netloc}"
    found = []
    seen  = set()

    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        full = urljoin(base, href)
        if urlparse(full).netloc != urlparse(source_url).netloc:
            continue
        if any(x in full for x in ("#", ".pdf", ".jpg", ".png", "javascript:", "mailto:")):
            continue
        path = urlparse(full).path
        if not re.search(r"/(news|article|story|post)/|/\d{6,}|[-a-z]{30,}", path):
            continue
        if full not in seen:
            seen.add(full)
            found.append(full)
        if len(found) >= max_links:
            break

    print(f"  Found {len(found)} article links")
    return found


# ── Ollama ────────────────────────────────────────────────────────────────────

def list_ollama_models(host: str) -> list[str]:
    try:
        r = requests.get(f"{host}/api/tags", timeout=5)
        r.raise_for_status()
        return [m["name"] for m in r.json().get("models", [])]
    except Exception:
        return []


def resolve_model(requested: str, available: list[str]) -> str:
    def base(m: str) -> str:
        return m.split(":")[0].lower()
    bases = [base(m) for m in available]
    if base(requested) in bases:
        return available[bases.index(base(requested))]
    return requested


def call_ollama(prompt: str, model: str, host: str, timeout: int = 120) -> str:
    payload = {
        "model": model,
        "prompt": prompt,
        "stream": False,
        "options": {"temperature": 0.1, "num_predict": 4096},
    }
    r = requests.post(f"{host}/api/generate", json=payload, timeout=timeout)
    r.raise_for_status()
    return r.json().get("response", "")


# ── Extraction prompt ─────────────────────────────────────────────────────────

INGEST_PROMPT = """\
You are extracting a structured story record from a healthcare news article.
Article source: "{source_name}"
Article URL: {url}

Extract ONE story object for a HIPAA and data breach intelligence database \
aimed at small healthcare practice owners (dental, medical, therapy, \
chiropractic — 1 to 50 employees).

Return ONLY a single valid JSON object. No array, no markdown, no explanation.
If the article is NOT about HIPAA, data breaches, patient data privacy, \
healthcare cybersecurity, or OCR enforcement, return exactly: {{}}

Schema:
{{
  "title": "Specific factual headline — name the entity and incident type if known",
  "category": "news" or "advice",
  "entity_size": "small" | "mid" | "large" | "vendor" | "unknown",
  "small_biz_relevance": "high" | "medium" | "low",
  "source": "{source_name}",
  "date": "publication date as Month YYYY or YYYY-MM-DD, or null",
  "summary": "2-3 sentences: what happened, consequence or penalty, and the takeaway for a small practice owner",
  "tags": ["3 to 5 tags — e.g. PHI breach, ransomware, OCR fine, business associate, phishing"],
  "penalty": "dollar amount string if enforcement action, else null",
  "affected_count": "number of patients/records as a string, or null",
  "url": "{url}"
}}

Relevance:
  high   — directly affects or could realistically happen to a small practice
  medium — large-org incident with a clear lesson for small practices
  low    — background context, less immediately actionable

--- ARTICLE CONTENT ---
{content}
--- END ---

Return ONLY the JSON object (or {{}} if not relevant):"""


def extract_story(url: str, source_name: str, model: str, host: str, verbose: bool) -> dict | None:
    """Fetch a URL, extract with Ollama, return story dict or None."""
    print(f"  Fetching : {url}", flush=True)
    content = fetch_page(url)

    if content.startswith("["):
        print(f"    ⚠ Skipping — {content[:100]}")
        return None

    if verbose:
        print(f"    Got {len(content)} chars")

    prompt = INGEST_PROMPT.format(
        source_name=source_name,
        url=url,
        content=content,
    )

    slug = url.rstrip("/").split("/")[-1][:60]
    print(f"  Analysing: {slug} ...", flush=True)

    try:
        response  = call_ollama(prompt, model, host, timeout=120)
        obj_start = response.find("{")
        obj_end   = response.rfind("}") + 1
        if obj_start == -1 or obj_end == 0:
            if verbose:
                print("    ⚠ No JSON object in response")
            return None

        story = json.loads(response[obj_start:obj_end])
        if not story or not story.get("title"):
            if verbose:
                print("    → Not relevant")
            return None

        if verbose:
            print(f"    → {story.get('title', '')[:70]}")
        return story

    except json.JSONDecodeError as e:
        if verbose:
            print(f"    ⚠ JSON parse error: {e}")
    except requests.exceptions.Timeout:
        print("    ⚠ Ollama timed out — try a smaller model")
    except Exception as e:
        if verbose:
            print(f"    ⚠ Error: {e}")
    return None


# ── Source name inference ─────────────────────────────────────────────────────

def infer_source_name(url: str) -> str:
    known = {
        "techtarget.com":       "TechTarget Healthtech Security",
        "healthtechsecurity":   "TechTarget Healthtech Security",
        "healthcaredive.com":   "Healthcare Dive",
        "hipaajournal.com":     "HIPAA Journal",
        "healthcareitnews.com": "Healthcare IT News",
        "govinfosecurity.com":  "GovInfoSecurity",
        "hhs.gov":              "HHS OCR",
        "cisa.gov":             "CISA",
        "oig.hhs.gov":          "HHS OIG",
        "justice.gov":          "DOJ",
        "eeoc.gov":             "EEOC",
        "osha.gov":             "OSHA",
        "maine.gov":            "Maine AG",
        "oag.ca.gov":           "California AG",
        "courtlistener.com":    "CourtListener",
    }
    netloc = urlparse(url).netloc.lower()
    for key, name in known.items():
        if key in netloc:
            return name
    return netloc.replace("www.", "").split(".")[0].title()


# ── URL collection ────────────────────────────────────────────────────────────

def read_urls_from_file(path: str) -> list[str]:
    """Read URLs from a plain-text file, one per line. Lines starting with # are comments."""
    urls = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                urls.append(line)
    return urls


def looks_like_file(value: str) -> bool:
    """Return True if a string looks like a local file path rather than a URL."""
    return not value.startswith("http") and (
        value.endswith(".txt") or os.path.isfile(value)
    )


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Ingest story URLs into the HIPAA breach intelligence database",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Specific story URLs
  python ingest.py --urls https://example.com/story1 https://example.com/story2

  # URLs from a text file (also works: --urls urls.txt)
  python ingest.py --file urls.txt

  # Crawl a source homepage for article links
  python ingest.py --source https://www.techtarget.com/healthtechsecurity

  # Dry run — extract and print without writing to Excel
  python ingest.py --urls https://example.com/story --dry-run

  # Custom model and output file
  python ingest.py --file urls.txt --model mistral --output my.xlsx
        """,
    )
    parser.add_argument("--urls",        nargs="+", metavar="URL_OR_FILE",
                        help="Story URLs to ingest, or a .txt file of URLs")
    parser.add_argument("--file",        metavar="FILE",
                        help="Plain-text file of URLs (one per line, # = comment)")
    parser.add_argument("--source",      metavar="URL",
                        help="Source index page to crawl for article links")
    parser.add_argument("--source-name", metavar="NAME",
                        help="Override the source label (default: inferred from domain)")
    parser.add_argument("--model",  "-m", default="llama3",
                        help="Ollama model name (default: llama3)")
    parser.add_argument("--host",         default="http://localhost:11434",
                        help="Ollama host URL")
    parser.add_argument("--output", "-o", default=EXCEL_FILE,
                        help=f"Excel output file (default: {EXCEL_FILE})")
    parser.add_argument("--max-links",   type=int, default=20,
                        help="Max links to extract from --source page (default: 20)")
    parser.add_argument("--delay",       type=float, default=1.0,
                        help="Seconds between URL fetches (default: 1.0)")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="Show detailed fetch and parse output")
    parser.add_argument("--dry-run",     action="store_true",
                        help="Extract and print without writing to Excel")
    parser.add_argument("--list-models", action="store_true",
                        help="List available Ollama models and exit")

    args = parser.parse_args()

    # ── List models ──
    if args.list_models:
        available = list_ollama_models(args.host)
        if available:
            print("\nAvailable Ollama models:")
            for m in available:
                print(f"  • {m}")
        else:
            print(f"\nNo models found — is Ollama running at {args.host}?")
        sys.exit(0)

    # ── Must have at least one input ──
    if not args.urls and not args.source and not args.file:
        parser.error("Provide at least one of --urls, --source, or --file")

    # ── Check Ollama ──
    available_models = list_ollama_models(args.host)
    if not available_models:
        print(f"\nError: Cannot reach Ollama at {args.host}", file=sys.stderr)
        print("Start Ollama with:  ollama serve", file=sys.stderr)
        sys.exit(1)

    model: str = resolve_model(args.model, available_models)
    if model == args.model and args.model not in available_models:
        print(f"\nWarning: model '{args.model}' not found locally.")
        print(f"Available: {', '.join(available_models)}")
        print(f"Pull it with:  ollama pull {args.model}\n")

    # ── Collect URLs ──
    all_urls: list[tuple[str, str]] = []  # (url, source_name)

    # --source crawl
    if args.source:
        sname = args.source_name or infer_source_name(args.source)
        for link in fetch_links_from_source(args.source, max_links=args.max_links):
            all_urls.append((link, sname))

    # --urls: each entry is either a URL or a file path
    if args.urls:
        for entry in args.urls:
            entry = entry.strip()
            if not entry:
                continue
            if looks_like_file(entry):
                # Treat as a file of URLs
                try:
                    file_urls = read_urls_from_file(entry)
                    print(f"  Reading {len(file_urls)} URLs from {entry}")
                    for url in file_urls:
                        sname = args.source_name or infer_source_name(url)
                        all_urls.append((url, sname))
                except FileNotFoundError:
                    print(f"  ⚠ File not found: {entry}")
            else:
                sname = args.source_name or infer_source_name(entry)
                all_urls.append((entry, sname))

    # --file
    if args.file:
        try:
            file_urls = read_urls_from_file(args.file)
            print(f"  Reading {len(file_urls)} URLs from {args.file}")
            for url in file_urls:
                sname = args.source_name or infer_source_name(url)
                all_urls.append((url, sname))
        except FileNotFoundError:
            print(f"Error: file not found: {args.file}", file=sys.stderr)
            sys.exit(1)

    # Deduplicate, preserve order
    seen_urls:   set[str]             = set()
    unique_urls: list[tuple[str,str]] = []
    for url, sn in all_urls:
        if url not in seen_urls:
            seen_urls.add(url)
            unique_urls.append((url, sn))

    if not unique_urls:
        print("No URLs to process.")
        sys.exit(0)

    action = "Appending to" if os.path.exists(args.output) else "Creating"
    print(f"\n{'='*60}")
    print(f"  URLs to process : {len(unique_urls)}")
    print(f"  Model           : {model}  |  Host: {args.host}")
    print(f"  Output          : {args.output}  ({action})")
    if args.dry_run:
        print(f"  Mode            : DRY RUN — nothing will be written")
    print(f"{'='*60}\n")

    # ── Process ──
    extracted: list[dict] = []
    skipped = 0

    for i, (url, source_name) in enumerate(unique_urls, 1):
        print(f"\n[{i}/{len(unique_urls)}] {source_name}", flush=True)
        story = extract_story(url, source_name, model, args.host, args.verbose)
        if story:
            extracted.append(story)
        else:
            skipped += 1
        if i < len(unique_urls):
            time.sleep(args.delay)

    # ── Summary ──
    print(f"\n{'='*60}")
    print(f"  Extracted : {len(extracted)} stories")
    print(f"  Skipped   : {skipped} (not relevant or fetch error)")
    print(f"{'='*60}\n")

    if not extracted:
        print("Nothing to write.")
        sys.exit(0)

    # ── Print terminal summary ──
    for i, story in enumerate(extracted, 1):
        cat       = story.get("category", "news")
        relevance = story.get("small_biz_relevance", "")
        badge     = {"high": "★★★", "medium": "★★ ", "low": "★  "}.get(relevance, "   ")
        label     = "[NEWS]  " if cat == "news" else "[ADVICE]"
        print(f"{i:>2}. {badge} {label}  {story.get('title', 'Untitled')}")
        print(f"     {story.get('source', '')}  |  {story.get('date', '')}")
        if story.get("penalty"):
            print(f"     Penalty: {story['penalty']}")
        if story.get("affected_count"):
            print(f"     Affected: {story['affected_count']}")
        summary = story.get("summary", "")
        if len(summary) > 200:
            summary = summary[:197] + "..."
        print(f"     {summary}")
        print(f"     {story.get('url', '')}")
        print()

    # ── Write Excel ──
    if args.dry_run:
        print("Dry run — Excel file not modified.")
    else:
        written = save_or_append_excel(extracted, model, path=args.output)
        print(f"\nDone. {written} new stories written to: {args.output}")


if __name__ == "__main__":
    main()