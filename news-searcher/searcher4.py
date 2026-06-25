#!/usr/bin/env python3
"""
Healthcare Compliance News Intelligence — Ollama Edition
Fetches raw content directly from US government enforcement sources,
then uses a local Ollama model to parse, classify, and summarise stories.

Dependencies: pip install requests beautifulsoup4 openpyxl
Requires:     Ollama running locally (default: http://localhost:11434)

On each run the script appends new stories to the top of the Excel file,
keeping entries in reverse-chronological order and deduplicating by title.
"""

import json
import argparse
import sys
import time
import re
import os
from datetime import datetime
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Constants ─────────────────────────────────────────────────────────────────

EXCEL_FILE  = "compliance_intelligence.xlsx"
SHEET_NAME  = "Stories"

DOMAIN_LABELS = {
    "all":    "All domains",
    "hipaa":  "HIPAA & Privacy",
    "cyber":  "Cybersecurity",
    "osha":   "OSHA & Safety",
    "fraud":  "Fraud & Abuse",
    "ethics": "Ethics & Licensing",
    "eeoc":   "Workplace & EEOC",
}

CATEGORY_LABELS = {
    "news":   "Breaking News",
    "advice": "Trend / Advice",
}

HEADERS = {
    # NOTE: a User-Agent that identifies as a bot (the previous value contained
    # "ComplianceIntelligenceBot") gets auto-blocked (403) by several .gov WAFs
    # even though a normal browser hitting the same URL works fine. Using a
    # realistic browser UA + the headers a real browser sends fixes most of
    # those 403s. This is just identification, not evasion of any auth/paywall.
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Connection": "keep-alive",
}


# ── Source definitions ────────────────────────────────────────────────────────

SOURCES = {
    "hipaa": [
        {
            "name": "HHS OCR Breach Portal",
            "url": "https://ocrportal.hhs.gov/ocr/breach/breach_report.jsf",
            "fallback_url": "https://www.hhs.gov/hipaa/for-professionals/breach-notification/breach-reporting/index.html",
            "strategy": "html",
            "domain": "hipaa",
            "category": "news",
        },
        {
            "name": "HHS OCR Resolution Agreements",
            "url": "https://www.hhs.gov/hipaa/for-professionals/compliance-enforcement/agreements/index.html",
            "fallback_url": "https://www.hhs.gov/hipaa/for-professionals/compliance-enforcement/enforcement-highlights/index.html",
            "strategy": "html",
            "domain": "hipaa",
            "category": "news",
        },
        {
            "name": "HHS OCR Press Releases",
            "url": "https://www.hhs.gov/about/news/index.html?q=hipaa&submit=Search",
            "fallback_url": "https://www.hhs.gov/hipaa/for-professionals/compliance-enforcement/agreements/index.html",
            "strategy": "html",
            "domain": "hipaa",
            "category": "news",
        },
        {
            # FIX: the old query used a bare "agencies[]=civil-rights-office" param.
            # The Federal Register API nests all filters under "conditions[...]",
            # so an unwrapped "agencies[]" is silently ignored — the call still
            # succeeds, but returns unfiltered "newest documents" instead of HIPAA-
            # related ones. There's also no separate agency slug for HHS OCR; OCR
            # filings are catalogued under the parent department. This version
            # filters by department + a HIPAA term search, which is the documented
            # correct format. Worth a quick --verbose check against your own run.
            "name": "Federal Register – HIPAA",
            "url": "https://www.federalregister.gov/api/v1/documents.json?conditions[term]=HIPAA&conditions[agencies][]=health-and-human-services-department&order=newest&per_page=10",
            "strategy": "json_api",
            "domain": "hipaa",
            "category": "advice",
        },
        {
            # FIX: the old URL pointed at a click-through consent page
            # ("/forms/acknowledgement.html"), not the breach list itself — a
            # plain GET can't get past that gate. This is the real breach-notices
            # landing page. Heads up: as of this writing Maine's own page states
            # their public-facing breach database is temporarily offline pending
            # a review, so this source may return little until that's restored;
            # it will resolve cleanly instead of erroring out either way.
            "name": "Maine AG Breach Database",
            "url": "https://www.maine.gov/ag/news-and-library/data-breach-notices",
            "strategy": "html",
            "domain": "hipaa",
            "category": "news",
        },
    ],
    "cyber": [
        {
            # FIX: this URL now 301-redirects to tech.hhs.gov, which is a
            # client-side rendered app — the HTML response has no real body
            # text for BeautifulSoup to extract (confirmed by fetching it
            # directly). HC3's content has moved to the HHS Cyber Gateway site
            # below, which still serves static, scrapable HTML. Its content is
            # mostly evergreen best-practice publications rather than dated
            # breaking news, so this is filed under "advice" now.
            "name": "HHS Cyber Gateway (HC3)",
            "url": "https://hhscyber.hhs.gov/resources-publications.html",
            "fallback_url": "https://hhscyber.hhs.gov/index.html",
            "strategy": "html",
            "domain": "cyber",
            "category": "advice",
        },
        {
            "name": "CISA Healthcare Advisories",
            "url": "https://www.cisa.gov/news-events/cybersecurity-advisories?f%5B0%5D=advisory_type%3A94",
            "strategy": "html",
            "domain": "cyber",
            "category": "news",
        },
        {
            "name": "DOJ Healthcare Data Breach Press Releases",
            "url": "https://www.justice.gov/news?q=healthcare+data+breach&op=Search",
            "strategy": "html",
            "domain": "cyber",
            "category": "news",
        },
    ],
    "osha": [
        {
            "name": "OSHA Press Releases",
            "url": "https://www.osha.gov/news/newsreleases",
            "strategy": "html",
            "domain": "osha",
            "category": "news",
        },
    ],
    "fraud": [
        {
            "name": "DOJ Health Care Fraud Unit",
            "url": "https://www.justice.gov/criminal/criminal-fraud/health-care-fraud-unit",
            "strategy": "html",
            "domain": "fraud",
            "category": "news",
        },
        {
            "name": "HHS OIG Fraud Alerts",
            "url": "https://oig.hhs.gov/compliance/alerts/",
            "strategy": "html",
            "domain": "fraud",
            "category": "news",
        },
        {
            "name": "HHS OIG Enforcement Actions",
            "url": "https://oig.hhs.gov/fraud/enforcement/",
            "strategy": "html",
            "domain": "fraud",
            "category": "news",
        },
    ],
    "ethics": [
        {
            "name": "Texas Medical Board Actions",
            "url": "https://www.tmb.state.tx.us/page/disciplinary-actions",
            "strategy": "html",
            "domain": "ethics",
            "category": "news",
        },
        {
            "name": "California Medical Board Actions",
            "url": "https://www.mbc.ca.gov/About/News/Press_Releases/",
            "strategy": "html",
            "domain": "ethics",
            "category": "news",
        },
    ],
    "eeoc": [
        {
            "name": "EEOC Healthcare Press Releases",
            "url": "https://www.eeoc.gov/newsroom/search?q=healthcare&type=pressrelease",
            "strategy": "html",
            "domain": "eeoc",
            "category": "news",
        },
    ],
}


# ── Source fetchers ───────────────────────────────────────────────────────────

def fetch_html(url: str, timeout: int = 15) -> str:
    """Fetch a page and return cleaned visible text (max 6000 chars).
    Falls back to verify=False on SSL errors (some .gov sites have cert issues).
    """
    import urllib3

    def _parse(r) -> str:
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
            tag.decompose()
        text = soup.get_text(separator="\n", strip=True)
        text = re.sub(r"\n{3,}", "\n\n", text)
        return text[:6000]

    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        return _parse(r)
    except requests.exceptions.SSLError:
        try:
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
            r = requests.get(url, headers=HEADERS, timeout=timeout, verify=False)
            r.raise_for_status()
            return _parse(r)
        except Exception as e2:
            return f"[SSL fallback also failed for {url}: {e2}]"
    except requests.exceptions.ConnectionError as e:
        return f"[Connection error for {url}: {e}]"
    except requests.exceptions.Timeout:
        return f"[Timeout fetching {url}]"
    except Exception as e:
        return f"[Could not fetch {url}: {e}]"


def fetch_json_api(url: str, timeout: int = 15) -> str:
    """Fetch a JSON API endpoint and return a readable summary."""
    try:
        r = requests.get(url, headers={**HEADERS, "Accept": "application/json"}, timeout=timeout)
        r.raise_for_status()
        data = r.json()
        results = data.get("results", data if isinstance(data, list) else [])
        lines = []
        for item in results[:15]:
            title    = item.get("title", "")
            date     = item.get("publication_date", item.get("date", ""))
            doc_type = item.get("type", "")
            abstract = item.get("abstract", "")[:300]
            lines.append(f"- [{date}] {doc_type}: {title}\n  {abstract}")
        return "\n".join(lines)[:5000]
    except Exception as e:
        return f"[Could not fetch JSON from {url}: {e}]"


def fetch_source(source: dict) -> str:
    """Fetch a source, trying fallback_url if primary fails."""
    if source.get("strategy") == "json_api":
        return fetch_json_api(source["url"])
    result = fetch_html(source["url"])
    if result.startswith("[") and source.get("fallback_url"):
        result2 = fetch_html(source["fallback_url"])
        if not result2.startswith("["):
            return result2
    return result


def is_fetch_error(text: str) -> bool:
    return text.startswith("[") and any(
        kw in text for kw in ("Could not fetch", "failed", "Timeout", "Connection error", "SSL")
    )


# ── Ollama interaction ────────────────────────────────────────────────────────

def list_ollama_models(host: str) -> list[str]:
    """Return list of locally available model names."""
    try:
        r = requests.get(f"{host}/api/tags", timeout=5)
        r.raise_for_status()
        return [m["name"] for m in r.json().get("models", [])]
    except Exception:
        return []


def resolve_model(requested: str, available: list[str]) -> str:
    """Match 'llama3.2' to 'llama3.2:latest' etc. Returns resolved name or original."""
    def base(m: str) -> str:
        return m.split(":")[0].lower()
    bases = [base(m) for m in available]
    if base(requested) in bases:
        return available[bases.index(base(requested))]
    return requested


def call_ollama(prompt: str, model: str, host: str, timeout: int = 180, num_predict: int = 1500) -> str:
    """Call Ollama via REST API and return the response text.

    FIX: the old hardcoded num_predict=4096 asks the model to generate far
    more tokens than a JSON extraction task needs, which is the main reason
    runs were timing out at 120s on CPU-only setups. 1500 tokens comfortably
    fits ~10 stories' worth of JSON. timeout is bumped to 180s as a buffer;
    both are now CLI-configurable (--timeout / --max-tokens) since the right
    values depend heavily on your hardware and which model you run.
    """
    payload = {
        "model": model,
        "prompt": prompt,
        "stream": False,
        "options": {"temperature": 0.1, "num_predict": num_predict},
    }
    r = requests.post(f"{host}/api/generate", json=payload, timeout=timeout)
    r.raise_for_status()
    return r.json().get("response", "")


def build_prompt(
    raw_content: str,
    source_name: str,
    domain: str,
    categories: list[str],
    max_results: int,
) -> str:
    if set(categories) == {"news", "advice"}:
        type_desc = "breaking news items OR trend/advice pieces"
    elif "news" in categories:
        type_desc = "breaking news items only"
    else:
        type_desc = "trend/advice pieces only"

    return f"""You are a healthcare compliance intelligence analyst. Below is raw content scraped from "{source_name}" — a US government or legal enforcement source.

Extract up to {max_results} distinct {type_desc} relevant to healthcare compliance teams. Focus on enforcement actions, breach reports, settlements, penalties, regulatory guidance, or cybersecurity threats.

Return ONLY a valid JSON array. No markdown, no explanation, no preamble. If there is nothing relevant, return [].

Each item schema:
{{
  "title": "Specific factual headline",
  "category": "news" or "advice",
  "domain": "{domain}",
  "source": "{source_name}",
  "date": "date or month/year string, or null",
  "summary": "2-3 sentences on what happened and why it matters",
  "tags": ["2 to 4 tags"],
  "penalty": "dollar amount string if enforcement action, else null",
  "url": "direct URL to the story or press release if found in the content, else null"
}}

--- RAW CONTENT START ---
{raw_content}
--- RAW CONTENT END ---

Return ONLY the JSON array:"""


# ── Core pipeline ─────────────────────────────────────────────────────────────

def fetch_stories(
    domain: str,
    categories: list[str],
    max_results: int,
    model: str,
    host: str,
    verbose: bool,
    ollama_timeout: int = 180,
    num_predict: int = 1500,
) -> list[dict]:
    """Fetch and extract stories from all relevant sources."""
    if domain == "all":
        sources = [s for group in SOURCES.values() for s in group]
    else:
        sources = SOURCES.get(domain, [])

    if not sources:
        raise ValueError(f"No sources configured for domain: {domain}")

    # Filter to requested categories; fall back to all if none match
    filtered = [s for s in sources if s["category"] in categories]
    sources  = filtered if filtered else sources

    all_stories: list[dict] = []
    per_source = max(2, max_results // max(len(sources), 1) + 1)

    for src in sources:
        print(f"  Fetching : {src['name']} ...", flush=True)
        raw = fetch_source(src)

        if is_fetch_error(raw):
            print(f"    ⚠ Skipping — {raw[:120]}", flush=True)
            continue

        if verbose:
            print(f"    Got {len(raw)} chars")

        prompt = build_prompt(raw, src["name"], src["domain"], categories, per_source)
        print(f"  Analysing: {src['name']} with {model} ...", flush=True)

        try:
            response  = call_ollama(prompt, model, host, timeout=ollama_timeout, num_predict=num_predict)
            start     = response.find("[")
            end       = response.rfind("]") + 1
            if start == -1 or end == 0:
                if verbose:
                    print("    ⚠ No JSON array in response")
                continue
            items = json.loads(response[start:end])
            items = [i for i in items if i.get("category") in categories]
            all_stories.extend(items)
            if verbose:
                print(f"    → {len(items)} stories extracted")
        except json.JSONDecodeError as e:
            if verbose:
                print(f"    ⚠ JSON parse error: {e}")
        except requests.exceptions.Timeout:
            print(f"    ⚠ Ollama timed out on {src['name']} — try a smaller model")
        except Exception as e:
            if verbose:
                print(f"    ⚠ Error: {e}")

        time.sleep(0.5)

    # Deduplicate by title and cap
    seen:   set[str]  = set()
    unique: list[dict] = []
    for story in all_stories:
        key = re.sub(r"[^a-z0-9]", "", story.get("title", "").lower())[:60]
        if key and key not in seen:
            seen.add(key)
            unique.append(story)

    return unique[:max_results]


# ── Excel output ──────────────────────────────────────────────────────────────

# Column layout: (header label, column width, story dict key)
COLUMNS = [
    ("Date",      16, "date"),
    ("Category",  13, "category"),
    ("Domain",    18, "domain"),
    ("Title",     52, "title"),
    ("Summary",   72, "summary"),
    ("Source",    28, "source"),
    ("Penalty",   16, "penalty"),
    ("Tags",      32, "tags"),
    ("URL",       52, "url"),
    ("Run Date",  18, "run_date"),
    ("Model",     20, "model"),
]

HEADER_BG  = "1F3864"
HEADER_FG  = "FFFFFF"
NEWS_BG    = "FFF2CC"
NEWS_ALT   = "FDE9A2"
ADVICE_BG  = "E2EFDA"
ADVICE_ALT = "D0E8C5"
BORDER_COL = "BFBFBF"


def _thin_border() -> Border:
    s = Side(style="thin", color=BORDER_COL)
    return Border(left=s, right=s, top=s, bottom=s)


def _write_header(ws) -> None:
    ws.row_dimensions[1].height = 22
    for col_idx, (label, width, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"


def _story_to_values(story: dict, run_date: str, model: str) -> list:
    tags = story.get("tags") or []
    tags_str = " · ".join(tags) if isinstance(tags, list) else str(tags)
    return [
        story.get("date") or "",
        (story.get("category") or "").title(),
        DOMAIN_LABELS.get(story.get("domain", ""), story.get("domain", "")),
        story.get("title") or "",
        story.get("summary") or "",
        story.get("source") or "",
        story.get("penalty") or "",
        tags_str,
        story.get("url") or "",
        run_date,
        model,
    ]


def _row_bg(category: str, row_num: int) -> str:
    even = (row_num % 2 == 0)
    if category.lower() == "news":
        return NEWS_ALT if even else NEWS_BG
    return ADVICE_ALT if even else ADVICE_BG


def _write_data_row(ws, row_num: int, values: list, category: str) -> None:
    bg      = _row_bg(category, row_num)
    url_col = next(i + 1 for i, (_, _, key) in enumerate(COLUMNS) if key == "url")
    wrap_cols = {
        i + 1 for i, (_, _, key) in enumerate(COLUMNS) if key in ("title", "summary", "tags")
    }

    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.border    = _thin_border()
        cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in wrap_cols))

        if col_idx == url_col and value:
            cell.hyperlink = value
            cell.font = Font(name="Arial", size=9, color="1155CC", underline="single")
        else:
            cell.font = Font(name="Arial", size=9)

    ws.row_dimensions[row_num].height = 48


def save_or_append_excel(stories: list[dict], model: str, path: str) -> int:
    """
    Write stories to an Excel file.
    - Creates the file if it does not exist.
    - Appends new stories at row 2 (below header) on subsequent runs,
      maintaining reverse-chronological order.
    - Deduplicates by title (case-insensitive).
    Returns the number of rows written.
    """
    run_date = datetime.now().strftime("%Y-%m-%d %H:%M")

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

        # Collect existing titles for deduplication (title is column 4, 0-based index 3)
        existing: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[3] if len(row) > 3 else None
            if val:
                existing.add(str(val).strip().lower())

        new_stories = [
            s for s in stories
            if str(s.get("title", "")).strip().lower() not in existing
        ]

        if not new_stories:
            print("  → No new stories (all duplicates of existing rows).")
            return 0

        ws.insert_rows(2, amount=len(new_stories))

        for i, story in enumerate(new_stories):
            row_num = 2 + i
            values  = _story_to_values(story, run_date, model)
            _write_data_row(ws, row_num, values, story.get("category", "news"))

        wb.save(path)
        print(f"  → Appended {len(new_stories)} new stories to: {path}")
        return len(new_stories)

    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.sheet_properties.tabColor = "1F3864"

        _write_header(ws)

        for i, story in enumerate(stories):
            row_num = 2 + i
            values  = _story_to_values(story, run_date, model)
            _write_data_row(ws, row_num, values, story.get("category", "news"))

        wb.save(path)
        print(f"  → Created: {path}  ({len(stories)} stories)")
        return len(stories)


# ── Terminal summary ──────────────────────────────────────────────────────────

def print_terminal(stories: list[dict], domain: str, categories: list[str], model: str) -> None:
    width = 72
    print()
    print("=" * width)
    print("  HEALTHCARE COMPLIANCE INTELLIGENCE REPORT")
    print(f"  Domain : {DOMAIN_LABELS.get(domain, domain)}")
    print(f"  Types  : {', '.join(CATEGORY_LABELS[c] for c in categories)}")
    print(f"  Model  : {model} (local Ollama)")
    print(f"  Run at : {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * width)

    news_count   = sum(1 for s in stories if s.get("category") == "news")
    advice_count = sum(1 for s in stories if s.get("category") == "advice")
    source_set   = {s.get("source", "Unknown") for s in stories}

    print(f"\n  {len(stories)} stories  |  {news_count} news  |  {advice_count} advice  |  {len(source_set)} sources\n")
    print("-" * width)

    for i, story in enumerate(stories, 1):
        cat     = story.get("category", "news")
        label   = "[NEWS]  " if cat == "news" else "[ADVICE]"
        title   = story.get("title", "Untitled")
        source  = story.get("source", "Unknown")
        date    = story.get("date") or ""
        penalty = story.get("penalty")
        summary = story.get("summary", "")
        tags    = story.get("tags", [])
        url     = story.get("url")

        print(f"\n{i:>2}. {label}  {title}")
        print(f"     Source: {source}" + (f"  |  {date}" if date else ""))
        if penalty:
            print(f"     Penalty: {penalty}")
        print()

        words, line, lines = summary.split(), [], []
        for w in words:
            if len(" ".join(line + [w])) > 66:
                lines.append("     " + " ".join(line))
                line = [w]
            else:
                line.append(w)
        if line:
            lines.append("     " + " ".join(line))
        print("\n".join(lines))

        if tags:
            tags_str = tags if isinstance(tags, str) else " · ".join(tags)
            print(f"\n     Tags: {tags_str}")
        if url:
            print(f"     URL : {url}")
        print()
        print("-" * width)
    print()


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Healthcare Compliance Intelligence — local Ollama edition",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python searcher.py
  python searcher.py --model mistral
  python searcher.py --domain cyber --type news
  python searcher.py --domain fraud --results 15
  python searcher.py --output my_report.xlsx
  python searcher.py --list-models
  python searcher.py --timeout 240 --max-tokens 1000   (slower hardware)

Domains: all, hipaa, cyber, osha, fraud, ethics, eeoc
Types:   news, advice, both (default: both)
Models:  any model pulled in Ollama (llama3, mistral, gemma3, etc.)
        """,
    )
    parser.add_argument("--domain",       "-d", choices=list(DOMAIN_LABELS.keys()), default="all")
    parser.add_argument("--type",         "-t", choices=["news", "advice", "both"],  default="both")
    parser.add_argument("--results",      "-n", type=int, default=10)
    parser.add_argument("--model",        "-m", default="llama3")
    parser.add_argument("--host",               default="http://localhost:11434")
    parser.add_argument("--output",       "-o", default=EXCEL_FILE)
    parser.add_argument("--verbose",      "-v", action="store_true")
    parser.add_argument("--list-models",        action="store_true")
    parser.add_argument("--timeout",            type=int, default=180,
                        help="Ollama request timeout in seconds (default: 180)")
    parser.add_argument("--max-tokens",         type=int, default=1500,
                        help="Ollama num_predict — lower if you keep timing out (default: 1500)")

    args = parser.parse_args()

    # ── List models and exit ──
    if args.list_models:
        available = list_ollama_models(args.host)
        if available:
            print("\nAvailable Ollama models:")
            for m in available:
                print(f"  • {m}")
        else:
            print(f"\nNo models found — is Ollama running at {args.host}?")
        sys.exit(0)

    # ── Check Ollama is reachable ──
    available_models = list_ollama_models(args.host)
    if not available_models:
        print(f"\nError: Cannot reach Ollama at {args.host}", file=sys.stderr)
        print("Make sure Ollama is running:  ollama serve", file=sys.stderr)
        sys.exit(1)

    # ── Resolve model name (e.g. "llama3.2" → "llama3.2:latest") ──
    resolved_model = resolve_model(args.model, available_models)
    if resolved_model == args.model and args.model not in available_models:
        print(f"\nWarning: model '{args.model}' not found locally.")
        print(f"Available: {', '.join(available_models)}")
        print(f"Pull it with:  ollama pull {args.model}\n")

    # ── Prepare run parameters ──
    categories: list[str] = ["news", "advice"] if args.type == "both" else [args.type]
    domain:     str       = args.domain
    model:      str       = resolved_model
    output:     str       = args.output
    action:     str       = "Appending to" if os.path.exists(output) else "Creating"

    print(f"\n{'='*60}")
    print(f"  Domain : {DOMAIN_LABELS[domain]}")
    print(f"  Types  : {', '.join(CATEGORY_LABELS[c] for c in categories)}")
    print(f"  Model  : {model}  |  Host: {args.host}")
    print(f"  Output : {output}  ({action})")
    print(f"{'='*60}\n")

    # ── Fetch stories ──
    try:
        stories: list[dict] = fetch_stories(
            domain=domain,
            categories=categories,
            max_results=args.results,
            model=model,
            host=args.host,
            verbose=args.verbose,
            ollama_timeout=args.timeout,
            num_predict=args.max_tokens,
        )
    except Exception as e:
        print(f"\nError: {e}", file=sys.stderr)
        sys.exit(1)

    if not stories:
        print("\nNo stories extracted. Try --verbose to debug, or a different domain.")
        sys.exit(0)

    # ── Write output ──
    save_or_append_excel(stories, model, path=output)
    print_terminal(stories, domain, categories, model)


if __name__ == "__main__":
    main()