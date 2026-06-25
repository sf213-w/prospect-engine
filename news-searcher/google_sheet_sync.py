#!/usr/bin/env python3
"""
Sync Excel → Google Sheets
Pushes rows from the compliance_intelligence.xlsx output (produced by
searcher.py) into a Google Sheet, so the data is viewable/shareable from
anywhere without opening the local Excel file.

Dependencies: pip install openpyxl gspread google-auth
Requires:     a Google Cloud service account with the Sheets API enabled,
              and the target Google Sheet shared with that service account's
              email as Editor. See SETUP below.

Default behavior is an INCREMENTAL sync: new rows (by Title, case-insensitive)
are inserted at the top of the Google Sheet tab, same as the Excel file's own
reverse-chronological layout. Use --full-resync to wipe the tab and rewrite it
to exactly match the Excel file (useful for the first run, or to recover from
manual edits made directly in Sheets).

SETUP (one-time):
  1. In Google Cloud Console, create/select a project and enable the
     "Google Sheets API".
  2. Create a Service Account, then generate a JSON key for it and save it
     as credentials.json next to this script (or pass --credentials).
  3. Open credentials.json and copy the "client_email" value.
  4. Create (or open) the target Google Sheet in your browser and share it
     with that client_email address, giving it Editor access.
  5. Run this script. It will create the "Stories" tab and header row on
     the first run if they don't already exist.
"""

import argparse
import os
import sys
from datetime import datetime

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook


# ── Constants ─────────────────────────────────────────────────────────────────

DEFAULT_EXCEL        = "compliance_intelligence.xlsx"
DEFAULT_EXCEL_SHEET  = "Stories"
DEFAULT_SHEET_NAME   = "Healthcare Compliance Intelligence"
DEFAULT_TAB          = "Stories"
DEFAULT_CREDENTIALS  = "credentials.json"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    # FIX: gc.open()/gc.create() look up the spreadsheet by NAME via the Drive
    # API, not the Sheets API — that needs its own scope. Without it, you get
    # "APIError: [403] Request had insufficient authentication scopes" even
    # though the sheet is correctly shared with the service account. This is
    # purely a client-side fix — no need to touch anything in Cloud Console
    # or regenerate credentials.json.
    "https://www.googleapis.com/auth/drive",
]

SYNCED_AT_HEADER = "Synced At"

HEADER_BG = {"red": 0.122, "green": 0.220, "blue": 0.392}   # matches the Excel navy (#1F3864)
HEADER_FG = {"red": 1, "green": 1, "blue": 1}


# ── Excel reading ─────────────────────────────────────────────────────────────

def read_excel_rows(path: str, sheet_name: str) -> tuple[list[str], list[list]]:
    """
    Read the source Excel file and return (headers, rows).
    - headers: the column labels from row 1, in order.
    - rows: every data row from row 2 onward, in the order they appear in the
      file (the Excel file already keeps newest stories at the top, so this
      list is newest-first).
    Missing cells come back as "" rather than None, to keep Sheets writes clean.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel file not found: {path}")

    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []

    headers = [str(h) if h is not None else "" for h in all_rows[0]]
    data_rows = [
        [("" if v is None else v) for v in row]
        for row in all_rows[1:]
        if any(v is not None and str(v).strip() != "" for v in row)   # skip fully blank rows
    ]
    return headers, data_rows


def _col_index(headers: list[str], name: str) -> int:
    """Case-insensitive header lookup. Raises a clear error if missing."""
    lowered = [h.strip().lower() for h in headers]
    target = name.strip().lower()
    if target not in lowered:
        raise ValueError(
            f"Expected a '{name}' column in the Excel sheet but didn't find one. "
            f"Found columns: {headers}"
        )
    return lowered.index(target)


# ── Google Sheets helpers ─────────────────────────────────────────────────────

def get_client(creds_path: str) -> gspread.Client:
    if not os.path.exists(creds_path):
        raise FileNotFoundError(
            f"Credentials file not found: {creds_path}\n"
            "See the SETUP steps in this script's docstring to create one."
        )
    creds = Credentials.from_service_account_file(creds_path, scopes=SCOPES)
    return gspread.authorize(creds)


def get_or_create_spreadsheet(gc: gspread.Client, name: str) -> gspread.Spreadsheet:
    try:
        return gc.open(name)
    except gspread.SpreadsheetNotFound:
        return gc.create(name)


def get_or_create_tab(sh: gspread.Spreadsheet, tab_name: str, headers: list[str]) -> gspread.Worksheet:
    """Open the tab if it exists, else create it with a styled header row."""
    try:
        ws = sh.worksheet(tab_name)
        return ws
    except gspread.WorksheetNotFound:
        pass

    ws = sh.add_worksheet(title=tab_name, rows=1000, cols=max(len(headers), 1))
    ws.update([headers], value_input_option="RAW")
    ws.format(
        f"A1:{gspread.utils.rowcol_to_a1(1, len(headers))}",
        {
            "backgroundColor": HEADER_BG,
            "textFormat": {"bold": True, "foregroundColor": HEADER_FG},
        },
    )
    ws.freeze(rows=1)
    try:
        ws.set_basic_filter()
    except Exception:
        pass  # cosmetic only — don't fail the sync over it

    _cleanup_default_tab(sh)
    return ws


def _cleanup_default_tab(sh: gspread.Spreadsheet) -> None:
    """Remove the blank default 'Sheet1' Google creates with every new spreadsheet."""
    try:
        default = sh.worksheet("Sheet1")
        if len(sh.worksheets()) > 1 and default.row_count and not default.get_all_values():
            sh.del_worksheet(default)
    except gspread.WorksheetNotFound:
        pass
    except Exception:
        pass


def get_existing_titles(ws: gspread.Worksheet, title_col_idx_1based: int) -> set[str]:
    values = ws.col_values(title_col_idx_1based)
    return {v.strip().lower() for v in values[1:] if v}  # skip header row


# ── Sync logic ────────────────────────────────────────────────────────────────

def sync_incremental(
    excel_path: str,
    excel_sheet: str,
    sheet_name: str,
    tab_name: str,
    creds_path: str,
    verbose: bool,
) -> tuple[int, int, str]:
    """
    Append only the Excel rows whose Title isn't already in the Google Sheet.
    Returns (new_rows_added, rows_already_present, spreadsheet_url).
    """
    headers, rows = read_excel_rows(excel_path, excel_sheet)
    if not headers:
        raise ValueError(f"No data found in '{excel_path}' (sheet: {excel_sheet}).")

    title_idx = _col_index(headers, "Title")
    sheet_headers = headers + [SYNCED_AT_HEADER]

    gc = get_client(creds_path)
    sh = get_or_create_spreadsheet(gc, sheet_name)
    ws = get_or_create_tab(sh, tab_name, sheet_headers)

    existing_titles = get_existing_titles(ws, title_idx + 1)
    synced_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    new_rows = []
    skipped = 0
    for row in rows:
        title = str(row[title_idx]).strip().lower() if len(row) > title_idx else ""
        if title and title in existing_titles:
            skipped += 1
            continue
        # Pad short rows so every row matches the header width before appending the timestamp
        padded = list(row) + [""] * (len(headers) - len(row))
        new_rows.append(padded + [synced_at])

    if new_rows:
        ws.insert_rows(new_rows, row=2, value_input_option="RAW")
        if verbose:
            print(f"  Inserted {len(new_rows)} new row(s) at the top of '{tab_name}'")

    return len(new_rows), skipped, sh.url


def sync_full(
    excel_path: str,
    excel_sheet: str,
    sheet_name: str,
    tab_name: str,
    creds_path: str,
    verbose: bool,
) -> tuple[int, str]:
    """
    Wipe the destination tab and rewrite it to exactly match the Excel file.
    Returns (rows_written, spreadsheet_url).
    """
    headers, rows = read_excel_rows(excel_path, excel_sheet)
    if not headers:
        raise ValueError(f"No data found in '{excel_path}' (sheet: {excel_sheet}).")

    sheet_headers = headers + [SYNCED_AT_HEADER]
    synced_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    gc = get_client(creds_path)
    sh = get_or_create_spreadsheet(gc, sheet_name)

    try:
        ws = sh.worksheet(tab_name)
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=tab_name, rows=max(len(rows) + 10, 100), cols=len(sheet_headers))

    body = [sheet_headers] + [
        list(row) + [""] * (len(headers) - len(row)) + [synced_at] for row in rows
    ]
    ws.update(body, value_input_option="RAW")
    ws.format(
        f"A1:{gspread.utils.rowcol_to_a1(1, len(sheet_headers))}",
        {
            "backgroundColor": HEADER_BG,
            "textFormat": {"bold": True, "foregroundColor": HEADER_FG},
        },
    )
    ws.freeze(rows=1)
    try:
        ws.set_basic_filter()
    except Exception:
        pass

    _cleanup_default_tab(sh)

    if verbose:
        print(f"  Rewrote '{tab_name}' with {len(rows)} row(s)")

    return len(rows), sh.url


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Sync the searcher.py Excel output to a Google Sheet.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python sync_to_sheets.py
  python sync_to_sheets.py --excel hipaa_breach_intelligence.xlsx
  python sync_to_sheets.py --sheet-name "Team Compliance Tracker"
  python sync_to_sheets.py --full-resync
  python sync_to_sheets.py --verbose
        """,
    )
    parser.add_argument("--excel",        "-e", default=DEFAULT_EXCEL,
                        help=f"Path to the source Excel file (default: {DEFAULT_EXCEL})")
    parser.add_argument("--excel-sheet",        default=DEFAULT_EXCEL_SHEET,
                        help=f"Worksheet name inside the Excel file (default: {DEFAULT_EXCEL_SHEET})")
    parser.add_argument("--sheet-name",         default=DEFAULT_SHEET_NAME,
                        help=f"Google Sheets file name — created if it doesn't exist (default: '{DEFAULT_SHEET_NAME}')")
    parser.add_argument("--tab",                default=DEFAULT_TAB,
                        help=f"Tab name inside the Google Sheet (default: {DEFAULT_TAB})")
    parser.add_argument("--credentials",  "-c", default=DEFAULT_CREDENTIALS,
                        help=f"Path to the service account JSON key (default: {DEFAULT_CREDENTIALS})")
    parser.add_argument("--full-resync",        action="store_true",
                        help="Wipe the destination tab and rewrite it to exactly match the Excel file")
    parser.add_argument("--verbose",      "-v", action="store_true")

    args = parser.parse_args()

    print(f"\n{'='*60}")
    print(f"  Source : {args.excel}  (sheet: {args.excel_sheet})")
    print(f"  Target : {args.sheet_name} → tab '{args.tab}'")
    print(f"  Mode   : {'Full resync (overwrite)' if args.full_resync else 'Incremental (append new only)'}")
    print(f"{'='*60}\n")

    try:
        if args.full_resync:
            written, url = sync_full(
                args.excel, args.excel_sheet, args.sheet_name, args.tab, args.credentials, args.verbose
            )
            print(f"  → Wrote {written} row(s) to the Google Sheet.")
        else:
            added, skipped, url = sync_incremental(
                args.excel, args.excel_sheet, args.sheet_name, args.tab, args.credentials, args.verbose
            )
            print(f"  → Added {added} new row(s)  |  {skipped} already present, skipped.")
        print(f"  → {url}\n")

    except FileNotFoundError as e:
        print(f"\nError: {e}", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"\nError: {e}", file=sys.stderr)
        sys.exit(1)
    except gspread.exceptions.APIError as e:
        print(f"\nGoogle Sheets API error: {e}", file=sys.stderr)
        print("Check that the service account has Editor access to the sheet.", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"\nUnexpected error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()