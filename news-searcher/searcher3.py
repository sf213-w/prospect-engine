#!/usr/bin/env python3
"""
HIPAA & Data Breach Intelligence for Small Healthcare Businesses
Fetches content from US government enforcement sources and uses a local
Ollama model to extract stories relevant to small practices (1-50 employees).

Dependencies: pip install requests beautifulsoup4 openpyxl
Requires:     Ollama running locally (default: http://localhost:11434)

On each run new stories are inserted at the top of the Excel file,
keeping entries in reverse-chronological order with duplicate detection.
"""

import json
import argparse
import sys
import time
import re
import os
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Constants ─────────────────────────────────────────────────────────────────

EXCEL_FILE = "hipaa_breach_intelligence.xlsx"
SHEET_NAME = "Stories"

CATEGORY_LABELS = {
    "news":   "Breaking News",
    "advice": "Trend / Advice",
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; ComplianceIntelligenceBot/1.0; research purposes)",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


# ── Sources ───────────────────────────────────────────────────────────────────
# Ordered by signal quality for HIPAA / data breach / small-practice relevance.
# priority 1 = highest value, run first.

SOURCES = [
    {
        "name": "HHS OCR Resolution Agreements & Penalties",
        "url": "https://www.hhs.gov/hipaa/for-professionals/compliance-enforcement/agreements/index.html",
        "fallback_url": "https://www.hhs.gov/hipaa/for-professionals/compliance-enforcement/enforcement-highlights/index.html",
        "strategy": "html",
        "category": "news",
        "priority": 1,
    },
    {
        "name": "HHS OCR Press Releases",
        "url": "https://www.hhs.gov/about/news/index.html?q=hipaa&submit=Search",
        "fallback_url": "https://www.hhs.gov/hipaa/for-professionals/compliance-enforcement/agreements/index.html",
        "strategy": "html",
        "category": "news",
        "priority": 1,
    },
    {
        "name": "HHS OCR Breach Portal",
        "url": "https://ocrportal.hhs.gov/ocr/breach/breach_report.jsf",
        "fallback_url": "https://www.hhs.gov/hipaa/for-professionals/breach-notification/breach-reporting/index.html",
        "strategy": "html",
        "category": "news",
        "priority": 1,
    },
    {
        "name": "Maine AG Breach Database",
        "url": "https://www.maine.gov/agviewer/content/ag/985235c7-cb95-819f-e996-b8f2f779ef3e/forms/acknowledgement.html",
        "strategy": "html",
        "category": "news",
        "priority": 2,
    },
    {
        "name": "California AG Data Breach List",
        "url": "https://oag.ca.gov/privacy/databreach/list",
        "strategy": "html",
        "category": "news",
        "priority": 2,
    },
    {
        "name": "CourtListener – HIPAA Class Actions",
        "url": "https://www.courtlistener.com/?type=r&q=HIPAA&order_by=score+desc&filed_after=2024-01-01",
        "strategy": "html",
        "category": "news",
        "priority": 2,
    },
    {
        "name": "HC3 Healthcare Cybersecurity Advisories",
        "url": "https://www.hhs.gov/about/agencies/asa/ocio/hc3/index.html",
        "strategy": "html",
        "category": "news",
        "priority": 2,
    },
    {
        "name": "Federal Register – HIPAA Rulemaking",
        "url": "https://www.federalregister.gov/api/v1/documents.json?agencies[]=civil-rights-office&order=newest&per_page=10",
        "strategy": "json_api",
        "category": "advice",
        "priority": 2,
    },
    {
        "name": "CISA Healthcare Cybersecurity Advisories",
        "url": "https://www.cisa.gov/news-events/cybersecurity-advisories?f%5B0%5D=advisory_type%3A94",
        "strategy": "html",
        "category": "news",
        "priority": 3,
    },
]


# ── Fetchers ──────────────────────────────────────────────────────────────────

def fetch_html(url: str, timeout: int = 15) -> str:
    """Fetch a page, strip chrome, return up to 6000 chars of visible text.
    Retries without SSL verification if the first attempt raises SSLError.
    """
    import urllib3

    def _extract(r) -> str:
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
            tag.decompose()
        text = soup.get_text(separator="\n", strip=True)
        return re.sub(r"\n{3,}", "\n\n", text)[:6000]

    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        return _extract(r)
    except requests.exceptions.SSLError:
        try:
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
            r = requests.get(url, headers=HEADERS, timeout=timeout, verify=False)
            r.raise_for_status()
            return _extract(r)
        except Exception as e:
            return f"[SSL fallback failed for {url}: {e}]"
    except requests.exceptions.ConnectionError as e:
        return f"[Connection error for {url}: {e}]"
    except requests.exceptions.Timeout:
        return f"[Timeout fetching {url}]"
    except Exception as e:
        return f"[Could not fetch {url}: {e}]"


def fetch_json_api(url: str, timeout: int = 15) -> str:
    """Fetch a JSON REST endpoint and return a readable text summary."""
    try:
        r = requests.get(url, headers={**HEADERS, "Accept": "application/json"}, timeout=timeout)
        r.raise_for_status()
        data    = r.json()
        results = data.get("results", data if isinstance(data, list) else [])
        lines   = []
        for item in results[:15]:
            title    = item.get("title", "")
            date     = item.get("publication_date", item.get("date", ""))
            doc_type = item.get("type", "")
            abstract = item.get("abstract", "")[:300]
            lines.append(f"- [{date}] {doc_type}: {title}\n  {abstract}")
        return "\n".join(lines)[:5000]
    except Exception as e:
        return f"[Could not fetch JSON from {url}: {e}]"


def fetch_source(source: dict) -> str:
    """Fetch a source, trying fallback_url if the primary URL fails."""
    if source.get("strategy") == "json_api":
        return fetch_json_api(source["url"])
    result = fetch_html(source["url"])
    if _is_error(result) and source.get("fallback_url"):
        fallback = fetch_html(source["fallback_url"])
        if not _is_error(fallback):
            return fallback
    return result


def _is_error(text: str) -> bool:
    return text.startswith("[") and any(
        kw in text for kw in ("Could not fetch", "failed", "Timeout", "Connection error", "SSL")
    )


# ── Ollama ────────────────────────────────────────────────────────────────────

def list_ollama_models(host: str) -> list[str]:
    try:
        r = requests.get(f"{host}/api/tags", timeout=5)
        r.raise_for_status()
        return [m["name"] for m in r.json().get("models", [])]
    except Exception:
        return []


def resolve_model(requested: str, available: list[str]) -> str:
    """Resolve 'llama3.2' → 'llama3.2:latest' etc."""
    def base(m: str) -> str:
        return m.split(":")[0].lower()
    bases = [base(m) for m in available]
    if base(requested) in bases:
        return available[bases.index(base(requested))]
    return requested


def call_ollama(prompt: str, model: str, host: str, timeout: int = 120) -> str:
    payload = {
        "model": model,
        "prompt": prompt,
        "stream": False,
        "options": {"temperature": 0.1, "num_predict": 4096},
    }
    r = requests.post(f"{host}/api/generate", json=payload, timeout=timeout)
    r.raise_for_status()
    return r.json().get("response", "")


def parse_json_array(text: str) -> list:
    """Extract and parse the first JSON array found in a string."""
    start = text.find("[")
    end   = text.rfind("]") + 1
    if start == -1 or end == 0:
        return []
    return json.loads(text[start:end])


# ── Prompts ───────────────────────────────────────────────────────────────────

EXTRACT_PROMPT = """\
You are a compliance intelligence analyst writing for owners of small healthcare businesses: \
independent medical practices, dental offices, mental health clinics, physical therapy offices, \
chiropractors, optometrists — typically 1 to 50 employees.

SOURCE: "{source_name}"

TASK: Extract up to {max_results} distinct stories from the content below.

INCLUDE — stories about:
- HIPAA violations, OCR enforcement actions, resolution agreements, civil monetary penalties
- PHI data breaches of any size (large-org breaches are useful as cautionary tales)
- Patient data exposed via hacking, ransomware, phishing, insider threats, lost/stolen devices
- Class action lawsuits filed after a healthcare data breach
- Business associate / vendor breaches (EHR vendors, billing companies, IT providers)
- New HIPAA rules, OCR guidance updates, or regulatory changes affecting small practices
- Cybersecurity threats targeting small healthcare providers or commonly used software/EHR platforms

EXCLUDE — do not extract:
- Large hospital system stories with no lesson for small practices
- Medicare/Medicaid billing fraud with no HIPAA or data-privacy angle
- OSHA citations, EEOC discrimination, medical licensing board actions
- Drug pricing, health insurance coverage policy, general health IT trends

FRAMING: Write each summary from the perspective of a practice owner. \
State what happened, the consequence or penalty, and what a small practice should do or watch for. \
Use plain language — avoid jargon.

Return ONLY a valid JSON array. No markdown, no preamble, no explanation. \
If nothing qualifies, return [].

Schema for each item:
{{
  "title": "Specific headline — name the entity and violation type if known",
  "category": "news" or "advice",
  "entity_size": "small" | "mid" | "large" | "vendor" | "unknown",
  "source": "{source_name}",
  "date": "date or month/year string, or null",
  "summary": "2-3 sentences: what happened, the penalty/consequence, and the takeaway for a small practice owner",
  "tags": ["2 to 5 tags — e.g. PHI breach, ransomware, OCR fine, business associate, phishing, lost device"],
  "penalty": "dollar amount string if enforcement action, else null",
  "affected_count": "number of patients or records affected as a string, or null",
  "url": "direct URL to the press release or source document if present in the content, else null"
}}

--- RAW CONTENT ---
{raw_content}
--- END ---

Return ONLY the JSON array:"""


RELEVANCE_FILTER_PROMPT = """\
You are a strict relevance filter for a HIPAA and data breach news feed \
aimed at owners of small healthcare practices (dental, medical, therapy, chiropractic — 1 to 50 employees).

Review the stories below and KEEP only those clearly about:
- HIPAA violations, OCR enforcement actions, PHI breaches, patient data exposure
- Ransomware, phishing, or cyberattacks on healthcare organisations
- Business associate / vendor data breaches affecting healthcare clients
- HIPAA regulatory changes or OCR guidance updates
- Class action lawsuits related to healthcare data breaches
- Cybersecurity threats to small healthcare practices or the software they use

REMOVE stories primarily about:
- Hospital system operations, mergers, financials unrelated to data/privacy
- Medicare/Medicaid billing fraud with no HIPAA angle
- OSHA, EEOC, medical licensing, drug pricing, general health policy

For each story that passes, set "small_biz_relevance":
  "high"   — directly affects or could realistically happen to a small practice
  "medium" — large-org incident with a clear lesson for small practices
  "low"    — relevant background context, less immediately actionable

Return ONLY a JSON array of passing stories, each with all original fields \
plus "small_biz_relevance". No markdown, no explanation.

Stories:
{stories_json}

Return ONLY the filtered JSON array:"""


# ── Pipeline ──────────────────────────────────────────────────────────────────

def extract_from_source(
    source: dict,
    categories: list[str],
    per_source: int,
    model: str,
    host: str,
    verbose: bool,
) -> list[dict]:
    """Fetch one source and ask the model to extract relevant stories."""
    print(f"  Fetching : {source['name']} ...", flush=True)
    raw = fetch_source(source)

    if _is_error(raw):
        print(f"    ⚠ Skipping — {raw[:120]}", flush=True)
        return []

    if verbose:
        print(f"    Got {len(raw)} chars")

    prompt = EXTRACT_PROMPT.format(
        source_name=source["name"],
        max_results=per_source,
        raw_content=raw,
    )
    print(f"  Analysing: {source['name']} ...", flush=True)

    try:
        response = call_ollama(prompt, model, host)
        items    = parse_json_array(response)
        items    = [i for i in items if i.get("category") in categories]
        if verbose:
            print(f"    → {len(items)} stories extracted")
        return items
    except json.JSONDecodeError as e:
        if verbose:
            print(f"    ⚠ JSON parse error: {e}")
    except requests.exceptions.Timeout:
        print(f"    ⚠ Ollama timed out on {source['name']} — try a smaller model")
    except Exception as e:
        if verbose:
            print(f"    ⚠ Error: {e}")
    return []


def relevance_filter(stories: list[dict], model: str, host: str, verbose: bool) -> list[dict]:
    """Second-pass model call: drop off-topic stories, score small-biz relevance."""
    if not stories:
        return []
    print(f"  Filtering {len(stories)} stories for small-biz HIPAA relevance ...", flush=True)
    prompt = RELEVANCE_FILTER_PROMPT.format(stories_json=json.dumps(stories, indent=2))
    try:
        response = call_ollama(prompt, model, host, timeout=120)
        filtered = parse_json_array(response)
        order    = {"high": 0, "medium": 1, "low": 2}
        filtered.sort(key=lambda s: order.get(s.get("small_biz_relevance", "low"), 2))
        if verbose:
            print(f"    → {len(filtered)} passed relevance filter")
        return filtered
    except Exception as e:
        if verbose:
            print(f"    ⚠ Filter error ({e}) — keeping unfiltered stories")
        return stories


def fetch_stories(
    categories: list[str],
    max_results: int,
    model: str,
    host: str,
    verbose: bool,
) -> list[dict]:
    """Full pipeline: fetch all sources, extract, deduplicate, relevance-filter."""
    sources         = sorted(SOURCES, key=lambda s: s.get("priority", 9))
    active_sources  = [s for s in sources if s["category"] in categories] or sources
    per_source      = max(3, max_results // max(len(active_sources), 1) + 2)

    all_stories: list[dict] = []
    for src in active_sources:
        items = extract_from_source(src, categories, per_source, model, host, verbose)
        all_stories.extend(items)
        time.sleep(0.5)

    # Deduplicate by normalised title
    seen:   set[str]   = set()
    unique: list[dict] = []
    for story in all_stories:
        key = re.sub(r"[^a-z0-9]", "", story.get("title", "").lower())[:60]
        if key and key not in seen:
            seen.add(key)
            unique.append(story)

    # Relevance filter + small-biz scoring
    unique = relevance_filter(unique, model, host, verbose)

    return unique[:max_results]


# ── Excel output ──────────────────────────────────────────────────────────────

# (header label, column width px, story dict key)
COLUMNS = [
    ("Date",               16, "date"),
    ("Category",           13, "category"),
    ("Relevance",          12, "small_biz_relevance"),
    ("Entity Size",        13, "entity_size"),
    ("Patients Affected",  18, "affected_count"),
    ("Title",              52, "title"),
    ("Summary",            72, "summary"),
    ("Source",             24, "source"),
    ("Penalty / Fine",     16, "penalty"),
    ("Tags",               32, "tags"),
    ("URL",                52, "url"),
    ("Run Date",           18, "run_date"),
    ("Model",              20, "model"),
]

# Title column index (1-based) — used for deduplication
_TITLE_COL = next(i + 1 for i, (_, _, k) in enumerate(COLUMNS) if k == "title")

HEADER_BG  = "1F3864"   # navy
HEADER_FG  = "FFFFFF"
HIGH_BG    = "FFF2CC"   # amber  — high relevance news
HIGH_ALT   = "FDE9A2"
MID_BG     = "E2EFDA"   # green  — medium relevance
MID_ALT    = "D0E8C5"
LOW_BG     = "EDF2F9"   # blue-grey — low / advice
LOW_ALT    = "DAE4F2"
BORDER_COL = "BFBFBF"


def _thin_border() -> Border:
    s = Side(style="thin", color=BORDER_COL)
    return Border(left=s, right=s, top=s, bottom=s)


def _write_header(ws) -> None:
    ws.row_dimensions[1].height = 22
    for col_idx, (label, width, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"


def _story_to_values(story: dict, run_date: str, model: str) -> list:
    tags = story.get("tags") or []
    return [
        story.get("date") or "",
        (story.get("category") or "").title(),
        (story.get("small_biz_relevance") or "").title(),
        (story.get("entity_size") or "").title(),
        story.get("affected_count") or "",
        story.get("title") or "",
        story.get("summary") or "",
        story.get("source") or "",
        story.get("penalty") or "",
        " · ".join(tags) if isinstance(tags, list) else str(tags),
        story.get("url") or "",
        run_date,
        model,
    ]


def _row_bg(story: dict, row_num: int) -> str:
    relevance = story.get("small_biz_relevance", "low")
    even      = (row_num % 2 == 0)
    if relevance == "high":
        return HIGH_ALT if even else HIGH_BG
    if relevance == "medium":
        return MID_ALT if even else MID_BG
    return LOW_ALT if even else LOW_BG


def _write_data_row(ws, row_num: int, values: list, story: dict) -> None:
    bg        = _row_bg(story, row_num)
    url_col   = next(i + 1 for i, (_, _, k) in enumerate(COLUMNS) if k == "url")
    wrap_cols = {i + 1 for i, (_, _, k) in enumerate(COLUMNS) if k in ("title", "summary", "tags")}

    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.border    = _thin_border()
        cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in wrap_cols))

        if col_idx == url_col and value:
            cell.hyperlink = value
            cell.font = Font(name="Arial", size=9, color="1155CC", underline="single")
        else:
            cell.font = Font(name="Arial", size=9)

    ws.row_dimensions[row_num].height = 52


def save_or_append_excel(stories: list[dict], model: str, path: str) -> int:
    """
    Create the Excel file on first run. On subsequent runs, insert new
    stories at row 2 (most recent at the top). Deduplicates by title.
    Returns number of rows written.
    """
    run_date  = datetime.now().strftime("%Y-%m-%d %H:%M")
    title_idx = _TITLE_COL - 1   # 0-based index for iter_rows values

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

        existing: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[title_idx] if len(row) > title_idx else None
            if val:
                existing.add(str(val).strip().lower())

        new_stories = [
            s for s in stories
            if str(s.get("title", "")).strip().lower() not in existing
        ]

        if not new_stories:
            print("  → No new stories (all duplicates of existing rows).")
            return 0

        ws.insert_rows(2, amount=len(new_stories))
        for i, story in enumerate(new_stories):
            row_num = 2 + i
            _write_data_row(ws, row_num, _story_to_values(story, run_date, model), story)

        wb.save(path)
        print(f"  → Appended {len(new_stories)} new stories to: {path}")
        return len(new_stories)

    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.sheet_properties.tabColor = "1F3864"

        _write_header(ws)
        for i, story in enumerate(stories):
            row_num = 2 + i
            _write_data_row(ws, row_num, _story_to_values(story, run_date, model), story)

        wb.save(path)
        print(f"  → Created: {path}  ({len(stories)} stories)")
        return len(stories)


# ── Terminal summary ──────────────────────────────────────────────────────────

def print_terminal(stories: list[dict], categories: list[str], model: str) -> None:
    width = 72
    print()
    print("=" * width)
    print("  HIPAA & BREACH INTELLIGENCE — SMALL BUSINESS FOCUS")
    print(f"  Types  : {', '.join(CATEGORY_LABELS[c] for c in categories)}")
    print(f"  Model  : {model} (local Ollama)")
    print(f"  Run at : {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * width)

    high   = sum(1 for s in stories if s.get("small_biz_relevance") == "high")
    medium = sum(1 for s in stories if s.get("small_biz_relevance") == "medium")
    news_c = sum(1 for s in stories if s.get("category") == "news")

    print(f"\n  {len(stories)} stories  |  {news_c} news  |  {high} high-relevance  |  {medium} medium\n")
    print("-" * width)

    for i, story in enumerate(stories, 1):
        cat       = story.get("category", "news")
        relevance = story.get("small_biz_relevance", "")
        rel_badge = {"high": "★★★", "medium": "★★ ", "low": "★  "}.get(relevance, "   ")
        label     = "[NEWS]  " if cat == "news" else "[ADVICE]"
        title     = story.get("title", "Untitled")
        source    = story.get("source", "Unknown")
        date      = story.get("date") or ""
        penalty   = story.get("penalty")
        affected  = story.get("affected_count")
        summary   = story.get("summary", "")
        tags      = story.get("tags", [])
        url       = story.get("url")

        print(f"\n{i:>2}. {rel_badge} {label}  {title}")
        print(f"     Source: {source}" + (f"  |  {date}" if date else ""))
        if penalty:
            print(f"     Penalty: {penalty}")
        if affected:
            print(f"     Affected: {affected} patients/records")
        print()

        words, line, lines = summary.split(), [], []
        for w in words:
            if len(" ".join(line + [w])) > 66:
                lines.append("     " + " ".join(line))
                line = [w]
            else:
                line.append(w)
        if line:
            lines.append("     " + " ".join(line))
        print("\n".join(lines))

        if tags:
            tags_str = tags if isinstance(tags, str) else " · ".join(tags)
            print(f"\n     Tags: {tags_str}")
        if url:
            print(f"     URL : {url}")
        print()
        print("-" * width)
    print()


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="HIPAA & Breach Intelligence for Small Healthcare Practices — Ollama edition",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python searcher.py
  python searcher.py --model mistral
  python searcher.py --type news --results 15
  python searcher.py --output weekly.xlsx
  python searcher.py --list-models
  python searcher.py --verbose

Types:   news, advice, both (default: both)
Models:  any model pulled in Ollama (llama3, mistral, gemma3, qwen2.5, etc.)
        """,
    )
    parser.add_argument("--type",         "-t", choices=["news", "advice", "both"], default="both",
                        help="Story category (default: both)")
    parser.add_argument("--results",      "-n", type=int, default=10,
                        help="Max stories to return (default: 10)")
    parser.add_argument("--model",        "-m", default="llama3",
                        help="Ollama model name (default: llama3)")
    parser.add_argument("--host",               default="http://localhost:11434",
                        help="Ollama host URL (default: http://localhost:11434)")
    parser.add_argument("--output",       "-o", default=EXCEL_FILE,
                        help=f"Excel output file (default: {EXCEL_FILE})")
    parser.add_argument("--verbose",      "-v", action="store_true",
                        help="Show detailed fetch and parse output")
    parser.add_argument("--list-models",        action="store_true",
                        help="List available Ollama models and exit")

    args = parser.parse_args()

    # ── List models ──
    if args.list_models:
        available = list_ollama_models(args.host)
        if available:
            print("\nAvailable Ollama models:")
            for m in available:
                print(f"  • {m}")
        else:
            print(f"\nNo models found — is Ollama running at {args.host}?")
        sys.exit(0)

    # ── Check Ollama ──
    available_models = list_ollama_models(args.host)
    if not available_models:
        print(f"\nError: Cannot reach Ollama at {args.host}", file=sys.stderr)
        print("Start Ollama with:  ollama serve", file=sys.stderr)
        sys.exit(1)

    # ── Resolve model ──
    model: str = resolve_model(args.model, available_models)
    if model == args.model and args.model not in available_models:
        print(f"\nWarning: model '{args.model}' not found locally.")
        print(f"Available: {', '.join(available_models)}")
        print(f"Pull it with:  ollama pull {args.model}\n")

    categories: list[str] = ["news", "advice"] if args.type == "both" else [args.type]
    output:     str       = args.output
    action:     str       = "Appending to" if os.path.exists(output) else "Creating"

    print(f"\n{'='*60}")
    print(f"  Focus  : HIPAA, Data Breaches, Small Healthcare Businesses")
    print(f"  Types  : {', '.join(CATEGORY_LABELS[c] for c in categories)}")
    print(f"  Model  : {model}  |  Host: {args.host}")
    print(f"  Output : {output}  ({action})")
    print(f"{'='*60}\n")

    # ── Run ──
    try:
        stories: list[dict] = fetch_stories(
            categories=categories,
            max_results=args.results,
            model=model,
            host=args.host,
            verbose=args.verbose,
        )
    except Exception as e:
        print(f"\nError: {e}", file=sys.stderr)
        sys.exit(1)

    if not stories:
        print("\nNo stories extracted. Try --verbose to debug, or run again.")
        sys.exit(0)

    save_or_append_excel(stories, model, path=output)
    print_terminal(stories, categories, model)


if __name__ == "__main__":
    main()
