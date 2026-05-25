"""
make_apollo_xlsx.py
-------------------
Reads a privacy_contacts_*.csv produced by npp_scraper.py and writes an
Apollo-ready Excel file alongside it.

Usage:
    python make_apollo_xlsx.py privacy_contacts_20260522_113957.csv

Apollo import columns produced:
    First Name        - contact first name (if found)
    Last Name         - contact last name (if found)
    Email             - primary provider-owned email
    Company           - provider_name from HHS data
    Company Website   - provider website
    Phone             - primary phone number (if found)
    Title             - inferred from context ("Privacy Officer" when known)
    Breach Date       - breach_submission_date (for outreach personalisation)
    Notes             - short summary of how the contact was found

Rows included:
    Only rows with at least one valid, provider-owned email address.
    Junk domains and generic local-parts are filtered out using the same
    two-stage logic as npp_scraper.py (domain blocklist + domain-match check).

Deduplication:
    When multiple rows exist for the same provider, the highest-scoring row
    is kept: +100 valid email, +10 phone, +5 site-crawl source, +2 name found.
"""

import csv
import os
import re
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Email validation (mirrors npp_scraper.py logic) ───────────────────────

JUNK_DOMAINS = {
    "va.gov", "department.va.gov", "hhs.gov", "ocr.hhs.gov",
    "dhs.gov", "cisa.dhs.gov", "cbp.dhs.gov", "fema.dhs.gov",
    "cms.gov", "ftc.gov", "justice.gov", "usdoj.gov",
    "jointcommission.org",
    "zoominfo.com", "signalhire.com", "leadiq.com", "npino.com",
    "npiprofile.com", "dnb.com", "pitchbook.com", "causeiq.com",
    "npienumerator.com",
    "pnc.com", "paubox.com", "molinahealthcare.com", "bluebeam.com",
    "mhnchicago.org", "eksm.com", "gialliance.com",
    "tiktok.com", "prnewswire.com", "nytimes.com", "businesslist.co.ke",
    "podchaser.com", "web.archive.org", "mass.gov",
    "classlawdc.com", "federmanlaw.com", "caffertyclobes.com",
    "pittmandutton.com", "simplyhired.com", "recovered.org",
    "medicalhomenetwork.org", "freeclinicdirectory.org",
}

JUNK_LOCAL_PARTS = {
    "privacyservice", "vhalonpublicaffairs", "vacovhaprivincoming",
    "complaint", "ocrcomplaint", "ocrmail", "ocr",
    "accommodationrequest", "fraud.foiarequests",
    "customerservice", "support",
}

_EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")


def _bare_domain(url):
    d = re.sub(r"https?://", "", url or "").split("/")[0].lower()
    return d.lstrip("www.")


def _first_valid_email(email_str, website="", source_url=""):
    """Return the first provider-owned email address, or '' if none."""
    if not email_str:
        return ""
    site_domain   = _bare_domain(website)
    source_domain = _bare_domain(source_url)
    for addr in re.split(r"[;,\s]+", email_str):
        addr = addr.strip()
        if not _EMAIL_RE.match(addr):
            continue
        local        = addr.split("@")[0].lower()
        email_domain = addr.split("@")[-1].lower().lstrip("www.")
        if email_domain in JUNK_DOMAINS:
            continue
        if any(email_domain.endswith("." + jd) for jd in JUNK_DOMAINS):
            continue
        if local in JUNK_LOCAL_PARTS:
            continue
        known = [d for d in (site_domain, source_domain) if d]
        if known:
            if not any(email_domain in kd or kd in email_domain for kd in known):
                continue
        return addr
    return ""


def _row_score(r):
    score = 0
    if _first_valid_email(r.get("emails", ""), r.get("website", ""), r.get("source_url", "")):
        score += 100
    if r.get("phones", "").strip():
        score += 10
    if r.get("found_via", "") == "site crawl":
        score += 5
    if r.get("first_name", "").strip():
        score += 2
    return score


def _first_phone(phones_str):
    """Return the first phone number from a semicolon-separated list."""
    if not phones_str:
        return ""
    parts = re.split(r"[;,]", phones_str)
    return parts[0].strip() if parts else ""


def _infer_title(context_snippet, found_via):
    """Best-effort role title from the page context."""
    if not context_snippet:
        return ""
    snippet_lower = context_snippet.lower()
    for title in [
        ("Chief Privacy Officer",     "chief privacy officer"),
        ("Privacy Officer",           "privacy officer"),
        ("HIPAA Officer",             "hipaa officer"),
        ("Compliance Officer",        "compliance officer"),
        ("Data Protection Officer",   "data protection officer"),
        ("Privacy Director",          "privacy director"),
        ("Privacy Coordinator",       "privacy coordinator"),
    ]:
        if title[1] in snippet_lower:
            return title[0]
    return "Privacy Contact"


# ── Styling helpers ────────────────────────────────────────────────────────

NAVY    = "1F3864"
BLUE    = "2E75B6"
WHITE   = "FFFFFF"
LGRAY   = "F2F2F2"
GOLD    = "C9A84C"
LBROWN  = "FFF3CD"


def _fill(hex6):
    return PatternFill("solid", fgColor=hex6)


def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr_font(color=WHITE):
    return Font(name="Arial", bold=True, size=10, color=color)


def _body_font(bold=False):
    return Font(name="Arial", size=10, bold=bold)


def _ctr():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _lft():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── Main builder ───────────────────────────────────────────────────────────

# Apollo's expected column names (exact strings for CSV/Excel import)
APOLLO_COLS = [
    ("First Name",       16),
    ("Last Name",        16),
    ("Email",            34),
    ("Company",          34),
    ("Company Website",  30),
    ("Phone",            18),
    ("Title",            24),
    ("Breach Date",      14),
    ("Notes",            44),
]


def build_apollo_xlsx(input_csv, output_xlsx=None):
    if output_xlsx is None:
        base        = os.path.splitext(input_csv)[0]
        output_xlsx = base + "_apollo.xlsx"

    # ── Load and deduplicate ───────────────────────────────────────────────
    with open(input_csv, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    best = {}
    for r in rows:
        key = r["provider_name"]
        if key not in best or _row_score(r) > _row_score(best[key]):
            best[key] = r

    apollo_rows = []
    for r in best.values():
        email = _first_valid_email(
            r.get("emails", ""),
            r.get("website", ""),
            r.get("source_url", ""),
        )
        if not email:
            continue  # only actionable contacts

        found_via = r.get("found_via", "")
        snippet   = r.get("context_snippet", "")
        note_parts = []
        if found_via == "site crawl":
            note_parts.append("contact found on provider privacy page")
        elif found_via == "web search":
            note_parts.append("contact found via web search")
        src = r.get("source_url", "")
        if src:
            note_parts.append(f"source: {src}")

        apollo_rows.append({
            "First Name":       r.get("first_name", "").strip(),
            "Last Name":        r.get("last_name", "").strip(),
            "Email":            email,
            "Company":          r.get("provider_name", "").strip(),
            "Company Website":  r.get("website", "").strip(),
            "Phone":            _first_phone(r.get("phones", "")),
            "Title":            _infer_title(snippet, found_via),
            "Breach Date":      r.get("breach_submission_date", "").strip(),
            "Notes":            "; ".join(note_parts),
        })

    # ── Build workbook ─────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Apollo Import"
    ws.sheet_properties.tabColor = GOLD
    ws.sheet_view.showGridLines   = False

    ncols = len(APOLLO_COLS)

    # Banner
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    banner = ws["A1"]
    banner.value     = (
        f"Apollo.io Import — Healthcare Privacy Contacts  |  "
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"{len(apollo_rows)} contacts"
    )
    banner.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
    banner.fill      = _fill(NAVY)
    banner.alignment = _ctr()
    ws.row_dimensions[1].height = 24

    # Column headers (row 2)
    for c, (h, _) in enumerate(APOLLO_COLS, 1):
        cell            = ws.cell(row=2, column=c, value=h)
        cell.font       = _hdr_font()
        cell.fill       = _fill(BLUE)
        cell.alignment  = _ctr()
        cell.border     = _thin()
    ws.row_dimensions[2].height = 28

    # Data rows
    for i, row in enumerate(apollo_rows, 1):
        r  = i + 2
        bg = WHITE if i % 2 else LGRAY
        for c, (col_name, _) in enumerate(APOLLO_COLS, 1):
            cell           = ws.cell(row=r, column=c, value=row[col_name])
            cell.font      = _body_font()
            cell.fill      = _fill(bg)
            cell.alignment = _lft()
            cell.border    = _thin()

    # Footer
    foot = len(apollo_rows) + 3
    ws.merge_cells(f"A{foot}:{get_column_letter(ncols)}{foot}")
    fc           = ws[f"A{foot}"]
    fc.value     = (
        f"{len(apollo_rows)} contacts ready for Apollo import  |  "
        f"Source: {os.path.basename(input_csv)}"
    )
    fc.font      = Font(name="Arial", bold=True, size=10, color=NAVY)
    fc.alignment = _lft()

    # Column widths
    for c, (_, w) in enumerate(APOLLO_COLS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A3"

    wb.save(output_xlsx)
    return output_xlsx, len(apollo_rows)


# ── CLI entry point ────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python make_apollo_xlsx.py <privacy_contacts_*.csv>")
        sys.exit(1)

    input_csv = sys.argv[1]
    if not os.path.exists(input_csv):
        print(f"Error: file not found — {input_csv}")
        sys.exit(1)

    out, count = build_apollo_xlsx(input_csv)
    print(f"✅ {count} contacts written to: {out}")
