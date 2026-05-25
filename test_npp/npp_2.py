import time
import re
import csv
import requests
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
from ddgs import DDGS
from urllib.parse import urlparse
from datetime import datetime
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

URL = "https://ocrportal.hhs.gov/ocr/breach/breach_report_hip.jsf"

TARGET_COUNT = 100
HHS_CSV_PATH = "hhs_breach_report.csv"  # updated by get_healthcare_providers()

CSV_FILE = f"privacy_contacts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
CSV_FIELDS = [
    "provider_name",
    "breach_submission_date",
    "first_name",
    "last_name",
    "website",
    "emails",
    "phones",
    "source_url",
    "context_snippet",
    "found_via",
    "date_scraped",
]


# -----------------------------
# CSV WRITER
# -----------------------------
def init_csv():
    with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
    print(f"📄 CSV initialized: {CSV_FILE}")


def write_csv_row(row: dict):
    with open(CSV_FILE, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writerow(row)


def write_not_found(provider_name, breach_date="", website=""):
    write_csv_row({
        "provider_name": provider_name,
        "breach_submission_date": breach_date,
        "first_name": "",
        "last_name": "",
        "website": website,
        "emails": "",
        "phones": "",
        "source_url": "",
        "context_snippet": "",
        "found_via": "not found",
        "date_scraped": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })


# -----------------------------
# DOWNLOAD HHS CSV EXPORT
# -----------------------------
def _find_export_button(page):
    """
    Try every known selector pattern for the PrimeFaces/JSF CSV export button.
    Returns the first matching ElementHandle, or None.

    The HHS portal renders this as an <a> or <span> with one of:
      - title="Export as CSV"
      - title containing "csv" or "export"
      - aria-label containing "csv" or "export"
      - inner text "CSV" or "Export"
      - an <img> whose src contains "csv" or "xls"
      - a PrimeFaces dataexporter icon class (e.g. ui-icon-arrowthickstop-1-s)
    """
    selectors = [
        # Exact title matches (most reliable when present)
        'a[title="Export as CSV"]',
        'a[title="CSV"]',
        'button[title="Export as CSV"]',
        'button[title="CSV"]',
        # Case-insensitive partial title via XPath
        '//a[contains(translate(@title,"csv","CSV"),"CSV")]',
        '//a[contains(translate(@title,"export","EXPORT"),"EXPORT")]',
        '//button[contains(translate(@title,"csv","CSV"),"CSV")]',
        # aria-label
        'a[aria-label*="CSV" i]',
        'a[aria-label*="export" i]',
        # Inner-text links/buttons
        'a:has-text("CSV")',
        'a:has-text("Export")',
        'button:has-text("CSV")',
        'button:has-text("Export")',
        # PrimeFaces dataexporter renders an <a> wrapping an <img>
        'a img[src*="csv"]',
        'a img[src*="xls"]',
        'a img[src*="export"]',
        # Spans/icons inside anchors
        'a .ui-icon-arrowthickstop-1-s',
        'a[id*="export"]',
        'a[id*="csv"]',
        'span[title*="CSV" i]',
        'span[title*="export" i]',
    ]

    for sel in selectors:
        try:
            if sel.startswith("//"):
                el = page.query_selector(f"xpath={sel}")
            else:
                el = page.query_selector(sel)
            if el and el.is_visible():
                print(f"  ✅ Export button found via selector: {sel!r}")
                return el
        except Exception:
            continue

    return None


def download_hhs_csv(download_path="hhs_breach_report.csv"):
    """
    Uses Playwright to load the HHS breach portal, waits for the table to
    fully render, then clicks the CSV export button and saves the download.
    Returns the path to the saved CSV file.
    """
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        print("Loading HHS portal...")
        page.goto(URL, timeout=60000)

        # Wait for the table to appear and populate
        print("Waiting for breach table to load...")
        try:
            page.wait_for_selector("tbody tr", timeout=30000)
        except Exception:
            page.screenshot(path="hhs_debug_no_table.png")
            raise RuntimeError(
                "Timed out waiting for the HHS table. "
                "The portal may be down for maintenance. "
                "A screenshot was saved to hhs_debug_no_table.png"
            )

        # Wait until actual row content is visible (portal lazy-loads)
        for _ in range(15):
            rows = page.query_selector_all("tbody tr")
            if rows and rows[0].inner_text().strip():
                break
            time.sleep(1)

        # Attempt to set rows-per-page to max so export covers everything
        try:
            rpp = page.query_selector(
                'select[id*="rpp"], select[id*="rows"], select[title*="Rows"]'
            )
            if rpp:
                rpp.select_option("100")
                page.wait_for_timeout(2000)
                page.wait_for_selector("tbody tr")
        except Exception:
            pass  # Not fatal

        # Locate the export button with multiple fallback selectors
        print("Locating CSV export button...")
        csv_btn = _find_export_button(page)

        if csv_btn is None:
            # Save a screenshot and dump all <a> tags to help debug
            page.screenshot(path="hhs_debug_no_button.png")
            links = []
            for a in page.query_selector_all("a"):
                try:
                    links.append({
                        "text":    a.inner_text().strip()[:50],
                        "title":   a.get_attribute("title") or "",
                        "id":      a.get_attribute("id") or "",
                        "class":   a.get_attribute("class") or "",
                        "onclick": (a.get_attribute("onclick") or "")[:80],
                    })
                except Exception:
                    pass
            print("  ⚠️  Could not find export button. All <a> tags on page:")
            for lnk in links:
                print(f"    {lnk}")
            browser.close()
            raise RuntimeError(
                "CSV export button not found on the HHS portal. "
                "A screenshot was saved to hhs_debug_no_button.png. "
                "Check the <a> tag dump above to identify the correct selector."
            )

        # Click and capture the file download
        print("Clicking CSV export button...")
        with page.expect_download(timeout=60000) as download_info:
            csv_btn.click()

        download = download_info.value
        download.save_as(download_path)
        print(f"✅ HHS CSV downloaded to: {download_path}")

        browser.close()
        return download_path


# -----------------------------
# PARSE HHS CSV → PROVIDER LIST
# -----------------------------
# Column indices in the HHS breach CSV export (0-based).
# The portal renders JSF component IDs as the first and eighth header cells,
# so we fall back to positional access for those two columns.
#
#  0 – Name of Covered Entity   (header is a JSF component ID, e.g. javax.faces…)
#  1 – State
#  2 – Covered Entity Type
#  3 – Individuals Affected
#  4 – Breach Submission Date
#  5 – Type of Breach
#  6 – Location of Breached Information
#  7 – Business Associate Present (also a JSF component ID)
#  8 – Web Description
_COL_NAME         = 0
_COL_ENTITY_TYPE  = 2
_COL_BREACH_DATE  = 4


def parse_hhs_csv(csv_path, target_count=TARGET_COUNT):
    """
    Reads the HHS breach CSV and returns up to `target_count` unique
    Healthcare Provider entries as (name, breach_submission_date) tuples.

    Column layout (positional — two headers are JSF component IDs):
      0  Name of Covered Entity
      1  State
      2  Covered Entity Type
      3  Individuals Affected
      4  Breach Submission Date
      5  Type of Breach
      6  Location of Breached Information
      7  Business Associate Present
      8  Web Description
    """
    providers = []
    seen = set()

    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)

        # Skip the header row
        headers = next(reader, None)
        if headers is None:
            print("  ⚠️  CSV appears empty.")
            return providers

        print(f"  CSV headers detected: {headers}")

        for row in reader:
            if len(row) <= max(_COL_NAME, _COL_ENTITY_TYPE, _COL_BREACH_DATE):
                continue  # malformed / short row

            name        = row[_COL_NAME].strip()
            entity_type = row[_COL_ENTITY_TYPE].strip()
            breach_date = row[_COL_BREACH_DATE].strip()

            if not name or not entity_type:
                continue

            if "Healthcare Provider" not in entity_type:
                continue

            if name not in seen:
                seen.add(name)
                providers.append((name, breach_date))
                print(f"  + {name} ({breach_date})")

            if len(providers) >= target_count:
                break

    return providers


# -----------------------------
# GET HEALTHCARE PROVIDERS (entry point)
# -----------------------------
def get_healthcare_providers(hhs_csv_path="hhs_breach_report.csv"):
    """
    Downloads the HHS breach CSV (or reuses an existing file) and returns
    up to TARGET_COUNT unique Healthcare Provider names.
    """
    global HHS_CSV_PATH
    HHS_CSV_PATH = hhs_csv_path

    # Re-use a previously downloaded file if it exists and is fresh (< 1 day old)
    if os.path.exists(hhs_csv_path):
        age_hours = (time.time() - os.path.getmtime(hhs_csv_path)) / 3600
        if age_hours < 24:
            print(f"♻️  Reusing existing HHS CSV ({age_hours:.1f}h old): {hhs_csv_path}")
        else:
            print("⏰ Existing HHS CSV is stale — re-downloading...")
            download_hhs_csv(hhs_csv_path)
    else:
        download_hhs_csv(hhs_csv_path)

    print(f"\n📋 Parsing providers from {hhs_csv_path}...")
    providers = parse_hhs_csv(hhs_csv_path)
    return providers


# -----------------------------
# DOMAIN FILTER
# -----------------------------
BAD_DOMAINS = [
    "linkedin.com", "facebook.com", "twitter.com", "instagram.com",
    "youtube.com", "healthgrades.com", "mapquest.com", "popupportal.com",
    "wikipedia.org", "yelp.com", "claimdepot.com", "databreach",
    "classaction.org", "topclassactions.com", "hipaajournal.com",
    "hhs.gov", "cms.gov", "bbb.org", "glassdoor.com", "indeed.com",
    "doximity.com", "vitals.com", "webmd.com", "zocdoc.com",
]

def is_valid_domain(url):
    return not any(bad in url for bad in BAD_DOMAINS)


# -----------------------------
# GET BASE URL (scheme + domain only)
# -----------------------------
def get_base_url(url):
    parsed = urlparse(url)
    return f"{parsed.scheme}://{parsed.netloc}"


# -----------------------------
# FIND OFFICIAL WEBSITE
# -----------------------------
def find_official_site(company):
    queries = [
        f'"{company}" official site privacy officer contact',
        f'"{company}" hospital clinic official website',
        f"{company} healthcare privacy HIPAA contact",
    ]

    for query in queries:
        try:
            with DDGS() as ddgs:
                results = list(ddgs.text(query, max_results=8))

            for r in results:
                url = r.get("href") or r.get("url")
                if not url:
                    continue
                if not is_valid_domain(url):
                    continue
                base = get_base_url(url)
                if base:
                    return base

        except Exception as e:
            print(f"  Search error: {e}")
            continue

    return None



# -----------------------------
# NAME EXTRACTION HELPERS
# -----------------------------

# Titles that signal a name is nearby
ROLE_TITLES = [
    "privacy officer", "chief privacy officer", "compliance officer",
    "data protection officer", "privacy director", "privacy manager",
    "privacy counsel", "hipaa officer", "hipaa privacy officer",
    "privacy contact", "privacy coordinator",
]

# Patterns that capture "FirstName LastName" near a role title.
# We look in both orders: "Name, Title" and "Title: Name" / "Title, Name"
_NAME_RE = re.compile(
    r"([A-Z][a-z]+(?:[\-'][A-Z][a-z]+)?)\s+([A-Z][a-z]+(?:[\-'][A-Z][a-z]+)?)"
)

def extract_name_from_snippet(snippet):
    """
    Attempt to pull a first/last name from a text snippet that is known to
    contain a privacy-role keyword.  Returns (first, last) or ("", "").

    Strategy:
      1. Find every "Word Word" capitalized pair within 120 chars of a role title.
      2. Skip pairs that are obviously not people names (all-caps abbreviations,
         known non-name words, the role title words themselves, etc.).
    """
    NON_NAME_WORDS = {
        "Privacy", "Officer", "Chief", "Compliance", "Data", "Protection",
        "Director", "Manager", "Counsel", "Coordinator", "Contact", "HIPAA",
        "Health", "Healthcare", "Medical", "Hospital", "Center", "Clinic",
        "Please", "Contact", "Notice", "Policy", "Information", "Rights",
        "Services", "Department", "Office", "Address", "Street", "Suite",
        "Phone", "Email", "Fax", "Mail", "Send", "Submit", "Form",
    }

    role_pattern = re.compile(
        r"(" + "|".join(re.escape(t) for t in ROLE_TITLES) + r")",
        re.IGNORECASE,
    )

    for role_match in role_pattern.finditer(snippet):
        # Look in a 200-char window around the role keyword
        start = max(0, role_match.start() - 150)
        end   = min(len(snippet), role_match.end() + 150)
        window = snippet[start:end]

        for m in _NAME_RE.finditer(window):
            first, last = m.group(1), m.group(2)
            if first in NON_NAME_WORDS or last in NON_NAME_WORDS:
                continue
            # Reject if either part is all-caps (likely an abbreviation)
            if first.isupper() or last.isupper():
                continue
            return first, last

    return "", ""


def infer_name_from_email(email):
    """
    Try to derive a name from an email local-part.
    Handles patterns like:
      john.smith@        → John, Smith
      jsmith@            → (no reliable first name)
      j.smith@           → (no reliable first name)
      johnsmith@         → (too ambiguous without a separator)
      john_smith@        → John, Smith
      john-smith@        → John, Smith
    Returns (first, last) or ("", "").
    """
    local = email.split("@")[0].lower()
    # Remove common prefixes like "privacy", "hipaa", "dpo", "contact", "info"
    generic = {"privacy", "hipaa", "dpo", "contact", "info", "admin",
               "compliance", "officer", "help", "support", "noreply", "no-reply"}
    if local in generic:
        return "", ""

    # Split on . _ -
    parts = re.split(r"[._\-]", local)
    parts = [p for p in parts if p]  # drop empty strings

    if len(parts) >= 2:
        first_part, last_part = parts[0], parts[-1]
        # Only accept if both parts look like real name segments (≥2 chars, alpha)
        if (len(first_part) >= 2 and first_part.isalpha() and
                len(last_part) >= 2 and last_part.isalpha()):
            return first_part.capitalize(), last_part.capitalize()

    return "", ""


def search_for_contact_name(provider, email):
    """
    Last-resort DDG search: look for the person's full name associated with
    this provider and email (or role title).
    Returns (first, last) or ("", "").
    """
    queries = [
        f'"{provider}" "privacy officer" name',
        f'"{provider}" HIPAA contact person name',
    ]
    for query in queries:
        try:
            with DDGS() as ddgs:
                results = list(ddgs.text(query, max_results=5))
            for r in results:
                snippet = r.get("body", "")
                first, last = extract_name_from_snippet(snippet)
                if first and last:
                    return first, last
        except Exception:
            continue
    return "", ""


def resolve_name(provider, emails, snippet):
    """
    Try all three strategies in order and return the first (first, last) hit.
      1. Extract from surrounding text snippet
      2. Infer from email local-part
      3. Web search as last resort
    """
    # Strategy 1: parse the context snippet
    first, last = extract_name_from_snippet(snippet)
    if first and last:
        return first, last

    # Strategy 2: infer from each email address
    for email in emails:
        first, last = infer_name_from_email(email)
        if first and last:
            return first, last

    # Strategy 3: targeted web search
    if emails:  # only bother if we have at least one email to anchor on
        first, last = search_for_contact_name(provider, emails[0])
        if first and last:
            return first, last

    return "", ""

# -----------------------------
# EXTRACT EMAILS FROM TEXT
# -----------------------------
def extract_emails(text):
    pattern = r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"
    return list(set(re.findall(pattern, text)))


# -----------------------------
# EXTRACT PRIVACY OFFICER INFO
# -----------------------------
def extract_privacy_contact(text, source_url, provider=""):
    results = []

    context_patterns = [
        r".{0,300}privacy officer.{0,300}",
        r".{0,300}compliance officer.{0,300}",
        r".{0,300}HIPAA.{0,300}",
        r".{0,300}data protection officer.{0,300}",
        r".{0,300}chief privacy.{0,300}",
    ]

    for pattern in context_patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            snippet = match.group()
            emails = extract_emails(snippet)
            phones = re.findall(r"\(?\d{3}\)?[\s\-\.]\d{3}[\s\-\.]\d{4}", snippet)
            first, last = extract_name_from_snippet(snippet)
            # If snippet alone didn't yield a name, try email inference
            if not (first and last):
                for email in emails:
                    first, last = infer_name_from_email(email)
                    if first and last:
                        break

            results.append({
                "snippet": snippet[:300].strip(),
                "emails": emails,
                "phones": phones,
                "source": source_url,
                "first_name": first,
                "last_name": last,
            })

    return results


# -----------------------------
# FETCH PAGE TEXT
# -----------------------------
def fetch_text(url):
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code != 200:
            return None
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "header"]):
            tag.decompose()
        return soup.get_text(" ", strip=True)
    except:
        return None


# -----------------------------
# SEARCH INSIDE SITE FOR PRIVACY PAGE
# -----------------------------
PRIVACY_PATHS = [
    "/privacy",
    "/privacy-policy",
    "/privacy-notice",
    "/hipaa",
    "/hipaa-notice",
    "/compliance",
    "/contact",
    "/contact-us",
    "/about/privacy",
    "/about/contact",
    "/patients/privacy",
    "/notices/privacy",
]

def find_privacy_contact(base_url, provider=""):
    all_contacts = []

    for path in PRIVACY_PATHS:
        url = base_url.rstrip("/") + path
        print(f"  Checking: {url}")
        text = fetch_text(url)

        if not text:
            continue

        contacts = extract_privacy_contact(text, url, provider)
        if contacts:
            all_contacts.extend(contacts)
            if any(c["emails"] for c in contacts):
                break

    return all_contacts


# -----------------------------
# SEARCH DDG FOR PRIVACY OFFICER EMAIL
# -----------------------------
def search_for_privacy_officer(company):
    """
    Returns (emails, source_url, first_name, last_name).
    """
    queries = [
        f'"{company}" "privacy officer" email contact',
        f'"{company}" HIPAA privacy notice contact email',
    ]

    for query in queries:
        try:
            with DDGS() as ddgs:
                results = list(ddgs.text(query, max_results=5))

            for r in results:
                url = r.get("href") or r.get("url")
                snippet_text = r.get("body", "")

                emails = extract_emails(snippet_text)
                if emails:
                    first, last = extract_name_from_snippet(snippet_text)
                    if not (first and last):
                        for email in emails:
                            first, last = infer_name_from_email(email)
                            if first and last:
                                break
                    return emails, url, first, last

                if url and is_valid_domain(url):
                    page_text = fetch_text(url)
                    if page_text:
                        contacts = extract_privacy_contact(page_text, url, company)
                        for c in contacts:
                            if c["emails"]:
                                return c["emails"], url, c["first_name"], c["last_name"]

        except Exception as e:
            print(f"  Search error: {e}")
            continue

    return [], None, "", ""



# ─────────────────────────────────────────────────────────────────────────────
# EXCEL REPORT BUILDER
# ─────────────────────────────────────────────────────────────────────────────

JUNK_DOMAINS = {
    "zoominfo.com","signalhire.com","va.gov","hhs.gov","dhs.gov",
    "jointcommission.org","pnc.com","molinahealthcare.com","bluebeam.com",
    "prnewswire.com","tiktok.com","leadiq.com","npino.com","mass.gov",
    "nytimes.com","businesslist.co.ke","causeiq.com","podchaser.com",
    "web.archive.org","pittmandutton.com","classlawdc.com","federmanlaw.com",
    "caffertyclobes.com","recovered.org","simplyhired.com",
}

_EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")

def _is_valid_email(email_str):
    if not email_str or not email_str.strip():
        return False
    for addr in re.split(r"[;,\s]+", email_str):
        addr = addr.strip()
        if _EMAIL_RE.match(addr):
            domain = addr.split("@")[-1].lower()
            if domain not in JUNK_DOMAINS:
                return True
    return False

def _scrape_summary(row):
    found_via = (row.get("found_via") or "").strip().lower()
    has_email = _is_valid_email(row.get("emails", ""))
    has_phone = bool((row.get("phones") or "").strip())
    has_site  = bool((row.get("website") or "").strip())
    has_name  = bool((row.get("first_name") or "").strip())

    if found_via == "not found":
        return "Could not find organization" if not has_site else "Found organization — no contact info retrieved"

    parts = []
    if has_site:  parts.append("org website found")
    if has_name:  parts.append("contact name identified")
    if has_email: parts.append("email found")
    if has_phone: parts.append("phone found")
    return ("Found: " + ", ".join(parts)) if parts else "Site found — no usable contact info"

def _row_score(r):
    score = 0
    if _is_valid_email(r.get("emails", "")): score += 100
    if r.get("phones", "").strip():           score += 10
    if r.get("found_via", "") == "site crawl": score += 5
    if r.get("first_name", "").strip():       score += 2
    return score

# Style constants
_NAVY      = "1F3864"; _STEEL     = "2E75B6"; _LGRAY    = "F2F2F2"
_WHITE     = "FFFFFF"; _GRN_HDR   = "375623"; _GRN_BG   = "E2EFDA"
_AMB_BG    = "FFF2CC"; _RED_HDR   = "843C0C"; _RED_BG   = "FCE4D6"

def _fill(h):   return PatternFill("solid", fgColor=h)
def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def _hfont(c=_WHITE): return Font(name="Arial", bold=True, color=c, size=10)
def _bfont():         return Font(name="Arial", size=10)
def _ctr():   return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _lft():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _style_hdr(ws, row, ncols, bg, fg=_WHITE):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _hfont(fg); cell.fill = _fill(bg)
        cell.alignment = _ctr(); cell.border = _thin()

def _style_data(ws, row, ncols, bg=_WHITE):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _bfont(); cell.fill = _fill(bg)
        cell.alignment = _lft(); cell.border = _thin()

def _banner(ws, text, ncols, bg, fg=_WHITE):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = text; c.font = Font(name="Arial", bold=True, size=13, color=fg)
    c.fill = _fill(bg); c.alignment = _ctr()
    ws.row_dimensions[1].height = 22

def _footer(ws, text, row, ncols, color):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws[f"A{row}"]
    c.value = text; c.font = Font(name="Arial", bold=True, size=10, color=color)
    c.alignment = _lft()

def _widths(ws, w_list):
    for i, w in enumerate(w_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def build_excel_report(hhs_csv_path, scraped_csv_path):
    """
    Build a three-sheet Excel workbook from the HHS source CSV and the
    scraped privacy-contacts CSV, then save it alongside the scraped CSV.

    Sheet 1 – HHS source data in full (with ID column)
    Sheet 2 – All processed / deduplicated results + search summary
    Sheet 3 – Actionable rows: valid email addresses only
    """
    # ── derive output path from scraped CSV name ───────────────────────────
    base     = os.path.splitext(scraped_csv_path)[0]
    out_path = base + ".xlsx"

    # ── load data ──────────────────────────────────────────────────────────
    hhs_rows, hhs_fields = [], []
    if os.path.exists(hhs_csv_path):
        with open(hhs_csv_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            raw_headers = next(reader, [])
            # Replace garbled JSF component IDs with readable labels
            CLEAN = {
                0: "Name of Covered Entity",
                7: "Business Associate Present",
            }
            hhs_fields = [CLEAN.get(i, h) for i, h in enumerate(raw_headers)]
            hhs_rows   = [dict(zip(hhs_fields, row)) for row in reader]

    with open(scraped_csv_path, newline="", encoding="utf-8-sig") as f:
        scraped_rows = list(csv.DictReader(f))

    # Deduplicate: one row per provider, keep highest-scoring
    best = {}
    for r in scraped_rows:
        key = r["provider_name"]
        if key not in best or _row_score(r) > _row_score(best[key]):
            best[key] = r
    deduped    = list(best.values())
    actionable = [r for r in deduped if _is_valid_email(r.get("emails", ""))]
    no_contact = [r for r in deduped if not _is_valid_email(r.get("emails", ""))]

    wb = Workbook()

    # ── Sheet 1: HHS source ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "1 - HHS Source Data"
    ws1.sheet_properties.tabColor = _STEEL

    hdr1 = ["ID"] + hhs_fields
    _banner(ws1, "HHS Breach Portal — Downloaded Source Data", len(hdr1), _NAVY)
    for c, h in enumerate(hdr1, 1):
        ws1.cell(row=2, column=c).value = h
    _style_hdr(ws1, 2, len(hdr1), _STEEL)
    ws1.row_dimensions[2].height = 30

    for i, row in enumerate(hhs_rows, 1):
        r  = i + 2
        bg = _WHITE if i % 2 else _LGRAY
        ws1.cell(row=r, column=1).value = i
        for c, field in enumerate(hhs_fields, 2):
            ws1.cell(row=r, column=c).value = row.get(field, "")
        _style_data(ws1, r, len(hdr1), bg)
        ws1.cell(row=r, column=1).alignment = _ctr()

    foot1 = len(hhs_rows) + 3
    _footer(ws1, f"Total records: {len(hhs_rows)}", foot1, len(hdr1), _NAVY)
    _widths(ws1, [5] + [max(14, min(len(f) + 2, 32)) for f in hhs_fields])
    ws1.freeze_panes = "A3"

    # ── Sheet 2: all processed results ────────────────────────────────────
    ws2 = wb.create_sheet("2 - Processed Results")
    ws2.sheet_properties.tabColor = _GRN_HDR

    S2 = [
        ("ID", 5), ("Provider Name", 32), ("Breach Date", 13),
        ("First Name", 12), ("Last Name", 12), ("Email(s)", 34),
        ("Phone(s)", 22), ("Website", 28), ("Source URL", 36),
        ("Found Via", 13), ("Search Summary", 42), ("Date Scraped", 18),
    ]
    _banner(ws2, "Breach Providers — Processed Contact Search Results", len(S2), _GRN_HDR)
    for c, (h, _) in enumerate(S2, 1):
        ws2.cell(row=2, column=c).value = h
    _style_hdr(ws2, 2, len(S2), _GRN_HDR)
    ws2.row_dimensions[2].height = 30

    for i, row in enumerate(deduped, 1):
        r  = i + 2
        bg = _GRN_BG if _is_valid_email(row.get("emails", "")) else _AMB_BG
        vals = [i, row.get("provider_name",""), row.get("breach_submission_date",""),
                row.get("first_name",""), row.get("last_name",""),
                row.get("emails",""), row.get("phones",""),
                row.get("website",""), row.get("source_url",""),
                row.get("found_via",""), _scrape_summary(row), row.get("date_scraped","")]
        for c, v in enumerate(vals, 1):
            ws2.cell(row=r, column=c).value = v
        _style_data(ws2, r, len(S2), bg)
        ws2.cell(row=r, column=1).alignment = _ctr()

    foot2 = len(deduped) + 3
    _footer(ws2,
        f"Unique providers: {len(deduped)}   |   With valid email: {len(actionable)}   |   No contact: {len(no_contact)}",
        foot2, len(S2), _GRN_HDR)
    _widths(ws2, [w for _, w in S2])
    ws2.freeze_panes = "A3"

    # ── Sheet 3: actionable — valid email only ─────────────────────────────
    ws3 = wb.create_sheet("3 - Valid Email Contacts")
    ws3.sheet_properties.tabColor = "C00000"

    S3 = S2  # same columns
    _banner(ws3, "Actionable Contacts — Valid Email Addresses Found", len(S3), _RED_HDR)
    for c, (h, _) in enumerate(S3, 1):
        ws3.cell(row=2, column=c).value = h
    _style_hdr(ws3, 2, len(S3), _RED_HDR)
    ws3.row_dimensions[2].height = 30

    for i, row in enumerate(actionable, 1):
        r  = i + 2
        bg = _WHITE if i % 2 else _RED_BG
        vals = [i, row.get("provider_name",""), row.get("breach_submission_date",""),
                row.get("first_name",""), row.get("last_name",""),
                row.get("emails",""), row.get("phones",""),
                row.get("website",""), row.get("source_url",""),
                row.get("found_via",""), _scrape_summary(row), row.get("date_scraped","")]
        for c, v in enumerate(vals, 1):
            ws3.cell(row=r, column=c).value = v
        _style_data(ws3, r, len(S3), bg)
        ws3.cell(row=r, column=1).alignment = _ctr()

    foot3 = len(actionable) + 3
    _footer(ws3, f"Actionable records with valid email: {len(actionable)}", foot3, len(S3), _RED_HDR)
    _widths(ws3, [w for _, w in S3])
    ws3.freeze_panes = "A3"

    wb.save(out_path)
    return out_path

# -----------------------------
# MAIN PIPELINE
# -----------------------------
def main():
    print("Starting full pipeline...\n")

    init_csv()

    # Download (or reuse) the HHS CSV export and extract provider names
    providers = get_healthcare_providers()

    if not providers:
        print("❌ No providers found")
        return

    print(f"\n✅ {len(providers)} providers found:")
    for name, breach_date in providers:
        print(f" - {name} ({breach_date})")

    print("\n🔍 Finding privacy officer contacts...\n")

    for provider, breach_date in providers:
        print(f"\n{'='*60}")
        print(f"  {provider}  [{breach_date}]")
        print(f"{'='*60}")

        # Step 1: Find the official site
        site = find_official_site(provider)
        if not site:
            print("  ❌ No official site found")
            write_not_found(provider, breach_date)
            continue

        print(f"  🌐 Website: {site}")

        found_anything = False
        scrape_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Step 2: Check known privacy-related paths on the site
        contacts = find_privacy_contact(site, provider)

        for c in contacts:
            if c["emails"] or c["phones"]:
                found_anything = True
                print(f"\n  ✅ Contact found at: {c['source']}")
                if c["emails"]:
                    print(f"     📧 Emails: {', '.join(c['emails'])}")
                if c["phones"]:
                    print(f"     📞 Phones: {', '.join(c['phones'])}")
                print(f"     📝 Context: ...{c['snippet'][:200]}...")

                # Last-resort web search for name if snippet + email both failed
                first, last = c["first_name"], c["last_name"]
                if not (first and last) and c["emails"]:
                    first, last = search_for_contact_name(provider, c["emails"][0])

                write_csv_row({
                    "provider_name": provider,
                    "breach_submission_date": breach_date,
                    "first_name": first,
                    "last_name": last,
                    "website": site,
                    "emails": "; ".join(c["emails"]),
                    "phones": "; ".join(c["phones"]),
                    "source_url": c["source"],
                    "context_snippet": c["snippet"][:300],
                    "found_via": "site crawl",
                    "date_scraped": scrape_time,
                })

        # Step 3: If nothing found on-site, try a targeted web search
        if not found_anything:
            print(f"\n  🔎 Trying targeted web search for privacy officer...")
            emails, source, first, last = search_for_privacy_officer(provider)
            if emails:
                print(f"  ✅ Found via web search!")
                print(f"     📧 Emails: {', '.join(emails)}")
                if first or last:
                    print(f"     👤 Name: {first} {last}".strip())
                print(f"     Source: {source}")

                write_csv_row({
                    "provider_name": provider,
                    "breach_submission_date": breach_date,
                    "first_name": first,
                    "last_name": last,
                    "website": site,
                    "emails": "; ".join(emails),
                    "phones": "",
                    "source_url": source,
                    "context_snippet": "",
                    "found_via": "web search",
                    "date_scraped": scrape_time,
                })
            else:
                print(f"  ❌ No privacy officer contact found")
                write_not_found(provider, breach_date, site)

    print(f"\n✅ Done. Results saved to: {CSV_FILE}")

    # Build Excel report automatically
    print("\n📊 Building Excel report...")
    try:
        xlsx_path = build_excel_report(HHS_CSV_PATH, CSV_FILE)
        print(f"📗 Excel report saved to: {xlsx_path}")
    except Exception as e:
        print(f"⚠️  Excel report failed: {e}")


if __name__ == "__main__":
    main()