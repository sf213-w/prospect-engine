import time
import re
import csv
import requests
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
from ddgs import DDGS
from urllib.parse import urlparse
from datetime import datetime
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

URL = "https://ocrportal.hhs.gov/ocr/breach/breach_report_hip.jsf"

TARGET_COUNT = 100
HHS_CSV_PATH = "hhs_breach_report.csv"  # updated by get_healthcare_providers()

CSV_FILE = f"privacy_contacts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
CSV_FIELDS = [
    "provider_name",
    "breach_submission_date",
    "first_name",
    "last_name",
    "website",
    "emails",
    "phones",
    "source_url",
    "context_snippet",
    "found_via",
    "date_scraped",
]


# -----------------------------
# CSV WRITER
# -----------------------------
def init_csv():
    with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
    print(f"📄 CSV initialized: {CSV_FILE}")


def write_csv_row(row: dict):
    with open(CSV_FILE, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writerow(row)


def write_not_found(provider_name, breach_date="", website=""):
    write_csv_row({
        "provider_name": provider_name,
        "breach_submission_date": breach_date,
        "first_name": "",
        "last_name": "",
        "website": website,
        "emails": "",
        "phones": "",
        "source_url": "",
        "context_snippet": "",
        "found_via": "not found",
        "date_scraped": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })


# -----------------------------
# DOWNLOAD HHS CSV EXPORT
# -----------------------------
def _find_export_button(page):
    """
    Try every known selector pattern for the PrimeFaces/JSF CSV export button.
    Returns the first matching ElementHandle, or None.

    The HHS portal renders this as an <a> or <span> with one of:
      - title="Export as CSV"
      - title containing "csv" or "export"
      - aria-label containing "csv" or "export"
      - inner text "CSV" or "Export"
      - an <img> whose src contains "csv" or "xls"
      - a PrimeFaces dataexporter icon class (e.g. ui-icon-arrowthickstop-1-s)
    """
    selectors = [
        # Exact title matches (most reliable when present)
        'a[title="Export as CSV"]',
        'a[title="CSV"]',
        'button[title="Export as CSV"]',
        'button[title="CSV"]',
        # Case-insensitive partial title via XPath
        '//a[contains(translate(@title,"csv","CSV"),"CSV")]',
        '//a[contains(translate(@title,"export","EXPORT"),"EXPORT")]',
        '//button[contains(translate(@title,"csv","CSV"),"CSV")]',
        # aria-label
        'a[aria-label*="CSV" i]',
        'a[aria-label*="export" i]',
        # Inner-text links/buttons
        'a:has-text("CSV")',
        'a:has-text("Export")',
        'button:has-text("CSV")',
        'button:has-text("Export")',
        # PrimeFaces dataexporter renders an <a> wrapping an <img>
        'a img[src*="csv"]',
        'a img[src*="xls"]',
        'a img[src*="export"]',
        # Spans/icons inside anchors
        'a .ui-icon-arrowthickstop-1-s',
        'a[id*="export"]',
        'a[id*="csv"]',
        'span[title*="CSV" i]',
        'span[title*="export" i]',
    ]

    for sel in selectors:
        try:
            if sel.startswith("//"):
                el = page.query_selector(f"xpath={sel}")
            else:
                el = page.query_selector(sel)
            if el and el.is_visible():
                print(f"  ✅ Export button found via selector: {sel!r}")
                return el
        except Exception:
            continue

    return None


def download_hhs_csv(download_path="hhs_breach_report.csv"):
    """
    Uses Playwright to load the HHS breach portal, waits for the table to
    fully render, then clicks the CSV export button and saves the download.
    Returns the path to the saved CSV file.
    """
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        print("Loading HHS portal...")
        page.goto(URL, timeout=60000)

        # Wait for the table to appear and populate
        print("Waiting for breach table to load...")
        try:
            page.wait_for_selector("tbody tr", timeout=30000)
        except Exception:
            page.screenshot(path="hhs_debug_no_table.png")
            raise RuntimeError(
                "Timed out waiting for the HHS table. "
                "The portal may be down for maintenance. "
                "A screenshot was saved to hhs_debug_no_table.png"
            )

        # Wait until actual row content is visible (portal lazy-loads)
        for _ in range(15):
            rows = page.query_selector_all("tbody tr")
            if rows and rows[0].inner_text().strip():
                break
            time.sleep(1)

        # Attempt to set rows-per-page to max so export covers everything
        try:
            rpp = page.query_selector(
                'select[id*="rpp"], select[id*="rows"], select[title*="Rows"]'
            )
            if rpp:
                rpp.select_option("100")
                page.wait_for_timeout(2000)
                page.wait_for_selector("tbody tr")
        except Exception:
            pass  # Not fatal

        # Locate the export button with multiple fallback selectors
        print("Locating CSV export button...")
        csv_btn = _find_export_button(page)

        if csv_btn is None:
            # Save a screenshot and dump all <a> tags to help debug
            page.screenshot(path="hhs_debug_no_button.png")
            links = []
            for a in page.query_selector_all("a"):
                try:
                    links.append({
                        "text":    a.inner_text().strip()[:50],
                        "title":   a.get_attribute("title") or "",
                        "id":      a.get_attribute("id") or "",
                        "class":   a.get_attribute("class") or "",
                        "onclick": (a.get_attribute("onclick") or "")[:80],
                    })
                except Exception:
                    pass
            print("  ⚠️  Could not find export button. All <a> tags on page:")
            for lnk in links:
                print(f"    {lnk}")
            browser.close()
            raise RuntimeError(
                "CSV export button not found on the HHS portal. "
                "A screenshot was saved to hhs_debug_no_button.png. "
                "Check the <a> tag dump above to identify the correct selector."
            )

        # Click and capture the file download
        print("Clicking CSV export button...")
        with page.expect_download(timeout=60000) as download_info:
            csv_btn.click()

        download = download_info.value
        download.save_as(download_path)
        print(f"✅ HHS CSV downloaded to: {download_path}")

        browser.close()
        return download_path


# -----------------------------
# PARSE HHS CSV → PROVIDER LIST
# -----------------------------
# Column indices in the HHS breach CSV export (0-based).
# The portal renders JSF component IDs as the first and eighth header cells,
# so we fall back to positional access for those two columns.
#
#  0 – Name of Covered Entity   (header is a JSF component ID, e.g. javax.faces…)
#  1 – State
#  2 – Covered Entity Type
#  3 – Individuals Affected
#  4 – Breach Submission Date
#  5 – Type of Breach
#  6 – Location of Breached Information
#  7 – Business Associate Present (also a JSF component ID)
#  8 – Web Description
_COL_NAME         = 0
_COL_ENTITY_TYPE  = 2
_COL_BREACH_DATE  = 4


def parse_hhs_csv(csv_path, target_count=TARGET_COUNT):
    """
    Reads the HHS breach CSV and returns up to `target_count` unique
    Healthcare Provider entries as (name, breach_submission_date) tuples.

    Column layout (positional — two headers are JSF component IDs):
      0  Name of Covered Entity
      1  State
      2  Covered Entity Type
      3  Individuals Affected
      4  Breach Submission Date
      5  Type of Breach
      6  Location of Breached Information
      7  Business Associate Present
      8  Web Description
    """
    providers = []
    seen = set()

    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)

        # Skip the header row
        headers = next(reader, None)
        if headers is None:
            print("  ⚠️  CSV appears empty.")
            return providers

        print(f"  CSV headers detected: {headers}")

        for row in reader:
            if len(row) <= max(_COL_NAME, _COL_ENTITY_TYPE, _COL_BREACH_DATE):
                continue  # malformed / short row

            name        = row[_COL_NAME].strip()
            entity_type = row[_COL_ENTITY_TYPE].strip()
            breach_date = row[_COL_BREACH_DATE].strip()

            if not name or not entity_type:
                continue

            if "Healthcare Provider" not in entity_type:
                continue

            if name not in seen:
                seen.add(name)
                providers.append((name, breach_date))
                print(f"  + {name} ({breach_date})")

            if len(providers) >= target_count:
                break

    return providers


# -----------------------------
# GET HEALTHCARE PROVIDERS (entry point)
# -----------------------------
def get_healthcare_providers(hhs_csv_path="hhs_breach_report.csv"):
    """
    Downloads the HHS breach CSV (or reuses an existing file) and returns
    up to TARGET_COUNT unique Healthcare Provider names.
    """
    global HHS_CSV_PATH
    HHS_CSV_PATH = hhs_csv_path

    # Re-use a previously downloaded file if it exists and is fresh (< 1 day old)
    if os.path.exists(hhs_csv_path):
        age_hours = (time.time() - os.path.getmtime(hhs_csv_path)) / 3600
        if age_hours < 24:
            print(f"♻️  Reusing existing HHS CSV ({age_hours:.1f}h old): {hhs_csv_path}")
        else:
            print("⏰ Existing HHS CSV is stale — re-downloading...")
            download_hhs_csv(hhs_csv_path)
    else:
        download_hhs_csv(hhs_csv_path)

    print(f"\n📋 Parsing providers from {hhs_csv_path}...")
    providers = parse_hhs_csv(hhs_csv_path)
    return providers


# -----------------------------
# DOMAIN FILTER
# -----------------------------
BAD_DOMAINS = [
    "linkedin.com", "facebook.com", "twitter.com", "instagram.com",
    "youtube.com", "healthgrades.com", "mapquest.com", "popupportal.com",
    "wikipedia.org", "yelp.com", "claimdepot.com", "databreach",
    "classaction.org", "topclassactions.com", "hipaajournal.com",
    "hhs.gov", "cms.gov", "bbb.org", "glassdoor.com", "indeed.com",
    "doximity.com", "vitals.com", "webmd.com", "zocdoc.com",
]

def is_valid_domain(url):
    return not any(bad in url for bad in BAD_DOMAINS)


# -----------------------------
# GET BASE URL (scheme + domain only)
# -----------------------------
def get_base_url(url):
    parsed = urlparse(url)
    return f"{parsed.scheme}://{parsed.netloc}"


# -----------------------------
# FIND OFFICIAL WEBSITE
# -----------------------------
def find_official_site(company):
    queries = [
        f'"{company}" official site privacy officer contact',
        f'"{company}" hospital clinic official website',
        f"{company} healthcare privacy HIPAA contact",
    ]

    for query in queries:
        try:
            with DDGS() as ddgs:
                results = list(ddgs.text(query, max_results=8))

            for r in results:
                url = r.get("href") or r.get("url")
                if not url:
                    continue
                if not is_valid_domain(url):
                    continue
                base = get_base_url(url)
                if base:
                    return base

        except Exception as e:
            print(f"  Search error: {e}")
            continue

    return None



# -----------------------------
# NAME EXTRACTION HELPERS
# -----------------------------

# Titles that signal a name is nearby
ROLE_TITLES = [
    "privacy officer", "chief privacy officer", "compliance officer",
    "data protection officer", "privacy director", "privacy manager",
    "privacy counsel", "hipaa officer", "hipaa privacy officer",
    "privacy contact", "privacy coordinator",
]

# Patterns that capture "FirstName LastName" near a role title.
# We look in both orders: "Name, Title" and "Title: Name" / "Title, Name"
_NAME_RE = re.compile(
    r"([A-Z][a-z]+(?:[\-'][A-Z][a-z]+)?)\s+([A-Z][a-z]+(?:[\-'][A-Z][a-z]+)?)"
)

def extract_name_from_snippet(snippet):
    """
    Attempt to pull a first/last name from a text snippet that is known to
    contain a privacy-role keyword.  Returns (first, last) or ("", "").

    Strategy:
      1. Find every "Word Word" capitalized pair within 120 chars of a role title.
      2. Skip pairs that are obviously not people names (all-caps abbreviations,
         known non-name words, the role title words themselves, etc.).
    """
    NON_NAME_WORDS = {
        "Privacy", "Officer", "Chief", "Compliance", "Data", "Protection",
        "Director", "Manager", "Counsel", "Coordinator", "Contact", "HIPAA",
        "Health", "Healthcare", "Medical", "Hospital", "Center", "Clinic",
        "Please", "Contact", "Notice", "Policy", "Information", "Rights",
        "Services", "Department", "Office", "Address", "Street", "Suite",
        "Phone", "Email", "Fax", "Mail", "Send", "Submit", "Form",
    }

    role_pattern = re.compile(
        r"(" + "|".join(re.escape(t) for t in ROLE_TITLES) + r")",
        re.IGNORECASE,
    )

    for role_match in role_pattern.finditer(snippet):
        # Look in a 200-char window around the role keyword
        start = max(0, role_match.start() - 150)
        end   = min(len(snippet), role_match.end() + 150)
        window = snippet[start:end]

        for m in _NAME_RE.finditer(window):
            first, last = m.group(1), m.group(2)
            if first in NON_NAME_WORDS or last in NON_NAME_WORDS:
                continue
            # Reject if either part is all-caps (likely an abbreviation)
            if first.isupper() or last.isupper():
                continue
            return first, last

    return "", ""


def infer_name_from_email(email):
    """
    Try to derive a name from an email local-part.
    Handles patterns like:
      john.smith@        → John, Smith
      jsmith@            → (no reliable first name)
      j.smith@           → (no reliable first name)
      johnsmith@         → (too ambiguous without a separator)
      john_smith@        → John, Smith
      john-smith@        → John, Smith
    Returns (first, last) or ("", "").
    """
    local = email.split("@")[0].lower()
    # Remove common prefixes like "privacy", "hipaa", "dpo", "contact", "info"
    generic = {"privacy", "hipaa", "dpo", "contact", "info", "admin",
               "compliance", "officer", "help", "support", "noreply", "no-reply"}
    if local in generic:
        return "", ""

    # Split on . _ -
    parts = re.split(r"[._\-]", local)
    parts = [p for p in parts if p]  # drop empty strings

    if len(parts) >= 2:
        first_part, last_part = parts[0], parts[-1]
        # Only accept if both parts look like real name segments (≥2 chars, alpha)
        if (len(first_part) >= 2 and first_part.isalpha() and
                len(last_part) >= 2 and last_part.isalpha()):
            return first_part.capitalize(), last_part.capitalize()

    return "", ""


def search_for_contact_name(provider, email):
    """
    Last-resort DDG search: look for the person's full name associated with
    this provider and email (or role title).
    Returns (first, last) or ("", "").
    """
    queries = [
        f'"{provider}" "privacy officer" name',
        f'"{provider}" HIPAA contact person name',
    ]
    for query in queries:
        try:
            with DDGS() as ddgs:
                results = list(ddgs.text(query, max_results=5))
            for r in results:
                snippet = r.get("body", "")
                first, last = extract_name_from_snippet(snippet)
                if first and last:
                    return first, last
        except Exception:
            continue
    return "", ""


def resolve_name(provider, emails, snippet):
    """
    Try all three strategies in order and return the first (first, last) hit.
      1. Extract from surrounding text snippet
      2. Infer from email local-part
      3. Web search as last resort
    """
    # Strategy 1: parse the context snippet
    first, last = extract_name_from_snippet(snippet)
    if first and last:
        return first, last

    # Strategy 2: infer from each email address
    for email in emails:
        first, last = infer_name_from_email(email)
        if first and last:
            return first, last

    # Strategy 3: targeted web search
    if emails:  # only bother if we have at least one email to anchor on
        first, last = search_for_contact_name(provider, emails[0])
        if first and last:
            return first, last

    return "", ""

# -----------------------------
# EXTRACT EMAILS FROM TEXT
# -----------------------------
def extract_emails(text):
    pattern = r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"
    return list(set(re.findall(pattern, text)))


# -----------------------------
# EXTRACT PRIVACY OFFICER INFO
# -----------------------------
def extract_privacy_contact(text, source_url, provider=""):
    results = []

    context_patterns = [
        r".{0,300}privacy officer.{0,300}",
        r".{0,300}compliance officer.{0,300}",
        r".{0,300}HIPAA.{0,300}",
        r".{0,300}data protection officer.{0,300}",
        r".{0,300}chief privacy.{0,300}",
    ]

    for pattern in context_patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            snippet = match.group()
            emails = extract_emails(snippet)
            phones = re.findall(r"\(?\d{3}\)?[\s\-\.]\d{3}[\s\-\.]\d{4}", snippet)
            first, last = extract_name_from_snippet(snippet)
            # If snippet alone didn't yield a name, try email inference
            if not (first and last):
                for email in emails:
                    first, last = infer_name_from_email(email)
                    if first and last:
                        break

            results.append({
                "snippet": snippet[:300].strip(),
                "emails": emails,
                "phones": phones,
                "source": source_url,
                "first_name": first,
                "last_name": last,
            })

    return results


# -----------------------------
# FETCH PAGE TEXT
# -----------------------------
def fetch_text(url):
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code != 200:
            return None
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "header"]):
            tag.decompose()
        return soup.get_text(" ", strip=True)
    except:
        return None


# -----------------------------
# SEARCH INSIDE SITE FOR PRIVACY PAGE
# -----------------------------
PRIVACY_PATHS = [
    "/privacy",
    "/privacy-policy",
    "/privacy-notice",
    "/hipaa",
    "/hipaa-notice",
    "/compliance",
    "/contact",
    "/contact-us",
    "/about/privacy",
    "/about/contact",
    "/patients/privacy",
    "/notices/privacy",
]

def find_privacy_contact(base_url, provider=""):
    all_contacts = []

    for path in PRIVACY_PATHS:
        url = base_url.rstrip("/") + path
        print(f"  Checking: {url}")
        text = fetch_text(url)

        if not text:
            continue

        contacts = extract_privacy_contact(text, url, provider)
        if contacts:
            all_contacts.extend(contacts)
            if any(c["emails"] for c in contacts):
                break

    return all_contacts


# -----------------------------
# SEARCH DDG FOR PRIVACY OFFICER EMAIL
# -----------------------------
def search_for_privacy_officer(company):
    """
    Returns (emails, source_url, first_name, last_name).
    """
    queries = [
        f'"{company}" "privacy officer" email contact',
        f'"{company}" HIPAA privacy notice contact email',
    ]

    for query in queries:
        try:
            with DDGS() as ddgs:
                results = list(ddgs.text(query, max_results=5))

            for r in results:
                url = r.get("href") or r.get("url")
                snippet_text = r.get("body", "")

                emails = extract_emails(snippet_text)
                if emails:
                    first, last = extract_name_from_snippet(snippet_text)
                    if not (first and last):
                        for email in emails:
                            first, last = infer_name_from_email(email)
                            if first and last:
                                break
                    return emails, url, first, last

                if url and is_valid_domain(url):
                    page_text = fetch_text(url)
                    if page_text:
                        contacts = extract_privacy_contact(page_text, url, company)
                        for c in contacts:
                            if c["emails"]:
                                return c["emails"], url, c["first_name"], c["last_name"]

        except Exception as e:
            print(f"  Search error: {e}")
            continue

    return [], None, "", ""



# ─────────────────────────────────────────────────────────────────────────────
# EXCEL REPORT BUILDER
# ─────────────────────────────────────────────────────────────────────────────

JUNK_DOMAINS = {
    "zoominfo.com","signalhire.com","va.gov","hhs.gov","dhs.gov",
    "jointcommission.org","pnc.com","molinahealthcare.com","bluebeam.com",
    "prnewswire.com","tiktok.com","leadiq.com","npino.com","mass.gov",
    "nytimes.com","businesslist.co.ke","causeiq.com","podchaser.com",
    "web.archive.org","pittmandutton.com","classlawdc.com","federmanlaw.com",
    "caffertyclobes.com","recovered.org","simplyhired.com",
}

_EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")

def _is_valid_email(email_str):
    if not email_str or not email_str.strip():
        return False
    for addr in re.split(r"[;,\s]+", email_str):
        addr = addr.strip()
        if _EMAIL_RE.match(addr):
            domain = addr.split("@")[-1].lower()
            if domain not in JUNK_DOMAINS:
                return True
    return False

def _scrape_summary(row):
    found_via = (row.get("found_via") or "").strip().lower()
    has_email = _is_valid_email(row.get("emails", ""))
    has_phone = bool((row.get("phones") or "").strip())
    has_site  = bool((row.get("website") or "").strip())
    has_name  = bool((row.get("first_name") or "").strip())

    if found_via == "not found":
        return "Could not find organization" if not has_site else "Found organization — no contact info retrieved"

    parts = []
    if has_site:  parts.append("org website found")
    if has_name:  parts.append("contact name identified")
    if has_email: parts.append("email found")
    if has_phone: parts.append("phone found")
    return ("Found: " + ", ".join(parts)) if parts else "Site found — no usable contact info"

def _row_score(r):
    score = 0
    if _is_valid_email(r.get("emails", "")): score += 100
    if r.get("phones", "").strip():           score += 10
    if r.get("found_via", "") == "site crawl": score += 5
    if r.get("first_name", "").strip():       score += 2
    return score

# Style constants
_NAVY      = "1F3864"; _STEEL     = "2E75B6"; _LGRAY    = "F2F2F2"
_WHITE     = "FFFFFF"; _GRN_HDR   = "375623"; _GRN_BG   = "E2EFDA"
_AMB_BG    = "FFF2CC"; _RED_HDR   = "843C0C"; _RED_BG   = "FCE4D6"

def _fill(h):   return PatternFill("solid", fgColor=h)
def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def _hfont(c=_WHITE): return Font(name="Arial", bold=True, color=c, size=10)
def _bfont():         return Font(name="Arial", size=10)
def _ctr():   return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _lft():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _style_hdr(ws, row, ncols, bg, fg=_WHITE):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _hfont(fg); cell.fill = _fill(bg)
        cell.alignment = _ctr(); cell.border = _thin()

def _style_data(ws, row, ncols, bg=_WHITE):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _bfont(); cell.fill = _fill(bg)
        cell.alignment = _lft(); cell.border = _thin()

def _banner(ws, text, ncols, bg, fg=_WHITE):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = text; c.font = Font(name="Arial", bold=True, size=13, color=fg)
    c.fill = _fill(bg); c.alignment = _ctr()
    ws.row_dimensions[1].height = 22

def _footer(ws, text, row, ncols, color):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws[f"A{row}"]
    c.value = text; c.font = Font(name="Arial", bold=True, size=10, color=color)
    c.alignment = _lft()

def _widths(ws, w_list):
    for i, w in enumerate(w_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _readme_row(ws, r, label, value, label_bg, label_fg=_WHITE, val_bg=_WHITE):
    """Write a two-cell label/value row on the README sheet."""
    lc = ws.cell(row=r, column=2, value=label)
    lc.font      = Font(name="Arial", bold=True, size=10, color=label_fg)
    lc.fill      = _fill(label_bg)
    lc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    lc.border    = _thin()

    vc = ws.cell(row=r, column=3, value=value)
    vc.font      = Font(name="Arial", size=10)
    vc.fill      = _fill(val_bg)
    vc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    vc.border    = _thin()
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    return r + 1


def _readme_section(ws, r, title, color):
    """Write a full-width section heading row."""
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    c = ws.cell(row=r, column=2, value=title)
    c.font      = Font(name="Arial", bold=True, size=11, color=_WHITE)
    c.fill      = _fill(color)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = _thin()
    ws.row_dimensions[r].height = 20
    return r + 1


def _readme_blank(ws, r):
    for col in range(2, 9):
        ws.cell(row=r, column=col).fill = _fill(_WHITE)
    return r + 1


def build_readme_sheet(wb, stats):
    """
    Insert a README worksheet as the first sheet in the workbook.
    `stats` is a dict with keys: hhs_total, scraped_raw, deduped,
    actionable, no_contact, junk_filtered, run_time.
    """
    ws = wb.create_sheet("README", 0)   # insert at position 0 (first tab)
    ws.sheet_properties.tabColor = _NAVY
    ws.sheet_view.showGridLines  = False
    ws.column_dimensions["A"].width = 2   # left margin
    ws.column_dimensions["B"].width = 30  # label column
    ws.column_dimensions["C"].width = 18
    for col in "DEFGH":
        ws.column_dimensions[col].width = 14

    # ── Title banner ───────────────────────────────────────────────────────
    ws.merge_cells("B1:H1")
    t = ws["B1"]
    t.value     = "HHS Breach Privacy Contact Scraper — Report Guide"
    t.font      = Font(name="Arial", bold=True, size=16, color=_WHITE)
    t.fill      = _fill(_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("B2:H2")
    sub = ws["B2"]
    sub.value     = f"Generated: {stats['run_time']}   |   Source: HHS OCR Breach Portal (ocrportal.hhs.gov)"
    sub.font      = Font(name="Arial", italic=True, size=9, color="595959")
    sub.fill      = _fill("D9E1F2")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    r = 4

    # ── SECTION 1: What this tool does ────────────────────────────────────
    r = _readme_section(ws, r, "▶  WHAT THIS TOOL DOES", _NAVY)
    ws.row_dimensions[r-1].height = 20
    r = _readme_row(ws, r, "Purpose",
        "Automatically scrapes the HHS HIPAA Breach Notification Portal, identifies Healthcare "
        "Providers who have reported data breaches, then searches each provider's website and "
        "the web to find a privacy officer contact email and phone number. Results are saved to "
        "a timestamped CSV and this Excel workbook.", _STEEL, val_bg="EEF3FB")
    ws.row_dimensions[r-1].height = 72
    r = _readme_row(ws, r, "Data Source",
        "U.S. Dept. of Health & Human Services - Office for Civil Rights (OCR) Breach Portal. ""URL: https://ocrportal.hhs.gov/ocr/breach/breach_report_hip.jsf ""The portal lists all HIPAA breaches affecting 500+ individuals reported in the last 24 months.",
        _STEEL, val_bg="EEF3FB")
    ws.row_dimensions[r-1].height = 54
    r = _readme_blank(ws, r)

    # ── SECTION 2: Run stats ───────────────────────────────────────────────
    r = _readme_section(ws, r, "▶  THIS RUN — AT A GLANCE", _STEEL)
    r = _readme_row(ws, r, "HHS records downloaded",
        f"{stats['hhs_total']} total breach records in the HHS source CSV (all entity types).",
        _STEEL, val_bg="EEF3FB")
    r = _readme_row(ws, r, "Healthcare Providers targeted",
        f"{stats['deduped']} unique Healthcare Providers extracted (Health Plans, Business Associates, "
        f"and duplicate provider names were excluded — see filtering rules below).",
        _STEEL, val_bg="EEF3FB")
    r = _readme_row(ws, r, "Raw scraper output rows",
        f"{stats['scraped_raw']} rows written to the CSV during scraping. One provider can produce "
        f"multiple rows when the scraper finds contacts on several privacy-related pages.",
        _STEEL, val_bg="EEF3FB")
    r = _readme_row(ws, r, "After deduplication",
        f"{stats['deduped']} unique providers — one representative row per provider kept (highest-scoring). "
        f"See 'Deduplication Rules' below.",
        _STEEL, val_bg="EEF3FB")
    r = _readme_row(ws, r, "Actionable (valid email found)",
        f"{stats['actionable']} providers — these appear in Sheet 4 and are ready to contact.",
        _STEEL, val_bg="EEF3FB")
    r = _readme_row(ws, r, "No contact found",
        f"{stats['no_contact']} providers — website may have been found but no usable email address "
        f"was retrieved. See Sheet 3 for detail.",
        _STEEL, val_bg="EEF3FB")
    r = _readme_row(ws, r, "Removed (junk email filtered)",
        f"{stats['junk_filtered']} rows were discarded because the email address found belonged to a "
        f"third-party domain (e.g. ZoomInfo, VA.gov, HHS.gov) — not the provider itself. "
        f"See 'Junk Domain Filter' below.",
        _STEEL, val_bg="EEF3FB")
    r = _readme_blank(ws, r)

    # ── SECTION 3: Sheet guide ─────────────────────────────────────────────
    r = _readme_section(ws, r, "▶  SHEET-BY-SHEET GUIDE", _GRN_HDR)
    r = _readme_row(ws, r, "Sheet 1 — README (this sheet)",
        "Overview, pipeline explanation, filtering rules, and run statistics.",
        _GRN_HDR, val_bg=_GRN_BG)
    r = _readme_row(ws, r, "Sheet 2 — HHS Source Data  🔵",
        "The complete, unmodified CSV downloaded from the HHS OCR Breach Portal. Every row is "
        "preserved, including Health Plans and Business Associates. An ID column (1, 2, 3…) is "
        "prepended for cross-referencing with the other sheets.",
        _GRN_HDR, val_bg=_GRN_BG)
    r = _readme_row(ws, r, "Sheet 3 — Processed Results  🟢/🟡",
        "One row per unique Healthcare Provider after deduplication. GREEN rows have a valid "
        "contact email. AMBER rows have a website or phone but no usable email. Includes the "
        "'Search Summary' column explaining what was and wasn't found for each provider.",
        _GRN_HDR, val_bg=_GRN_BG)
    r = _readme_row(ws, r, "Sheet 4 — Valid Email Contacts  🔴",
        "Actionable subset: only the providers from Sheet 3 where a real, provider-owned email "
        "address was found. This is the list CJ works from to send outreach emails.",
        _GRN_HDR, val_bg=_GRN_BG)
    r = _readme_blank(ws, r)

    # ── SECTION 4: Pipeline steps ─────────────────────────────────────────
    r = _readme_section(ws, r, "▶  PIPELINE STEPS (in order)", _NAVY)
    steps = [
        ("Step 1 — Download HHS CSV",
         "Playwright (a headless browser) loads the HHS portal, waits for the breach table to "
         "fully render, then clicks the CSV export button. The downloaded file is cached as "
         "'hhs_breach_report.csv' and reused for 24 hours to avoid hammering the portal."),
        ("Step 2 — Filter to Healthcare Providers",
         "Reads the CSV positionally (column 0 = provider name, col 2 = entity type, col 4 = "
         "breach date). Only rows where Covered Entity Type contains 'Healthcare Provider' are "
         "kept. Health Plans, Business Associates, and unknown types are skipped."),
        ("Step 3 — Find official website",
         "DuckDuckGo is searched with three queries per provider: (a) official site + privacy "
         "officer, (b) hospital/clinic + official website, (c) healthcare + HIPAA contact. The "
         "first result from a non-blacklisted domain is used as the provider's website."),
        ("Step 4 — Crawl known privacy paths",
         "Up to 12 URL paths are checked on the provider's site in order: /privacy, "
         "/privacy-policy, /hipaa, /hipaa-notice, /compliance, /contact, /contact-us, "
         "/about/privacy, /about/contact, /patients/privacy, /notices/privacy. Crawling stops "
         "at the first path that returns an email address."),
        ("Step 5 — Extract contact info from page text",
         "Each page is stripped of scripts/styles/nav, then searched with 5 regex patterns "
         "around keywords: 'privacy officer', 'compliance officer', 'HIPAA', 'data protection "
         "officer', 'chief privacy'. Emails and phone numbers are extracted from a 300-char "
         "window around each match."),
        ("Step 6 — Attempt to identify contact name",
         "Three strategies are tried in order: (a) find a 'FirstName LastName' capitalized pair "
         "within 150 chars of a role-title keyword on the page; (b) infer from the email "
         "local-part if it contains a dot/underscore separator (e.g. john.smith@ → John Smith); "
         "(c) run a targeted DuckDuckGo search for '[Provider] privacy officer name'."),
        ("Step 7 — Web search fallback",
         "If no email was found by crawling, two DDG queries are run: '[Provider] privacy "
         "officer email contact' and '[Provider] HIPAA privacy notice contact email'. Results "
         "are checked for emails in the snippet; if none, the linked page is fetched and "
         "re-parsed."),
        ("Step 8 — Write CSV row",
         "One row is written per contact found (a provider with multiple privacy pages may "
         "produce multiple rows at this stage). Providers with no contact info at all are "
         "written with found_via='not found'."),
        ("Step 9 — Build Excel report",
         "After all providers are processed, this workbook is generated automatically. The raw "
         "CSV rows are deduplicated (one row per provider, best score kept), split into "
         "actionable vs. no-contact, and laid out across Sheets 2–4."),
    ]
    for label, detail in steps:
        r = _readme_row(ws, r, label, detail, _NAVY, val_bg="EEF3FB")
        ws.row_dimensions[r-1].height = 60
    r = _readme_blank(ws, r)

    # ── SECTION 5: Why entries are removed ───────────────────────────────
    r = _readme_section(ws, r, "▶  WHY ENTRIES ARE REMOVED OR DOWNGRADED", _RED_HDR)
    removals = [
        ("Not a Healthcare Provider",
         "The HHS CSV includes Health Plans, Business Associates, and hybrid entities. Only rows "
         "where 'Covered Entity Type' is exactly 'Healthcare Provider' are processed. All others "
         "are silently skipped at Step 2."),
        ("Duplicate provider name",
         "When a provider appears multiple times in the HHS CSV (e.g., reported two separate "
         "breaches), only the first occurrence is processed. Duplicates are tracked via a 'seen' "
         "set and skipped."),
        ("Official website not found",
         "If all three DuckDuckGo queries for a provider return only blacklisted domains (social "
         "media, aggregators, news sites), no website is recorded and the provider is written as "
         "'not found'. This happens most often for very small, local, or recently-closed practices."),
        ("Website on blacklisted domain",
         "Even if a search result is returned, the URL is rejected if its domain appears in "
         "BAD_DOMAINS: LinkedIn, Facebook, Twitter, Instagram, YouTube, Healthgrades, Yelp, "
         "Wikipedia, WebMD, ZocDoc, Doximity, Vitals, HHS.gov, CMS.gov, Glassdoor, Indeed, "
         "BBB, MapQuest, classaction sites, and HIPAA news blogs."),
        ("No contact info on privacy pages",
         "The 12 standard privacy-path URLs are fetched and parsed. If none of them contain "
         "text matching the privacy-officer keywords, or if matches exist but contain no email "
         "or phone number, the provider falls through to the web-search fallback (Step 7)."),
        ("Junk email domain filtered (Sheet 4 only)",
         "Even when an email is found, it is excluded from Sheet 4 (actionable) if its domain "
         "is in JUNK_DOMAINS. These are addresses that the scraper mistakenly retrieved from "
         "third-party pages instead of the provider's own site. Current junk list: zoominfo.com, "
         "signalhire.com, va.gov, hhs.gov, dhs.gov, jointcommission.org, pnc.com, "
         "molinahealthcare.com, bluebeam.com, prnewswire.com, tiktok.com, leadiq.com, "
         "npino.com, mass.gov, nytimes.com, businesslist.co.ke, causeiq.com, podchaser.com, "
         "web.archive.org, simplyhired.com, classlawdc.com, federmanlaw.com, recovered.org."),
        ("Multiple raw rows collapsed to one",
         "When the scraper finds contacts on several privacy-related pages for the same provider, "
         "each page produces a separate CSV row. During Excel generation, these are collapsed to "
         "a single row per provider using a scoring system: +100 for valid email, +10 for phone, "
         "+5 for site-crawl source (preferred over web-search), +2 for a name identified. The "
         "highest-scoring row wins."),
        ("Contact name left blank",
         "Name extraction is best-effort. If no capitalized 'First Last' pair appears near a "
         "role-title keyword, the email local-part has no separator, and the DDG name search "
         "returns nothing useful, first_name and last_name are left empty. The row is still "
         "kept — a blank name does not remove an entry."),
    ]
    for label, detail in removals:
        r = _readme_row(ws, r, label, detail, _RED_HDR, val_bg=_RED_BG)
        ws.row_dimensions[r-1].height = 72
    r = _readme_blank(ws, r)

    # ── SECTION 6: Search Summary values ─────────────────────────────────
    r = _readme_section(ws, r, "▶  'SEARCH SUMMARY' COLUMN — POSSIBLE VALUES", _STEEL)
    summaries = [
        ("Found: org website found, email found, phone found",
         "Best case. Official site identified, valid email address and phone number both retrieved."),
        ("Found: org website found, email found",
         "Official site found and a valid email retrieved. No phone number on the privacy page."),
        ("Found: org website found, contact name identified, email found",
         "As above, plus the contact person's name was identified from the page or email address."),
        ("Found: org website found, phone found",
         "Site found and phone retrieved, but no email address — amber row, not in Sheet 4."),
        ("Site found — no usable contact info",
         "A valid domain was found but no email or phone appeared on any of the 12 checked paths, "
         "and the web-search fallback also returned nothing."),
        ("Found organization — no contact info retrieved",
         "A website was identified, but all 12 privacy-path fetches failed (404s, timeouts, or "
         "pages with no matching keywords), and the fallback search also found nothing."),
        ("Could not find organization",
         "DuckDuckGo returned only blacklisted domains for all three queries. The provider has "
         "no usable web presence in the search index."),
    ]
    for label, detail in summaries:
        r = _readme_row(ws, r, label, detail, _STEEL, val_bg="EEF3FB")
        ws.row_dimensions[r-1].height = 48

    # ── set consistent row heights for label rows ─────────────────────────
    for row_idx in range(1, r):
        if ws.row_dimensions[row_idx].height is None or ws.row_dimensions[row_idx].height == 0:
            ws.row_dimensions[row_idx].height = 48

    return ws


def build_excel_report(hhs_csv_path, scraped_csv_path):
    """
    Build a four-sheet Excel workbook from the HHS source CSV and the
    scraped privacy-contacts CSV, then save it alongside the scraped CSV.

    Sheet 1 – README: pipeline guide, filtering rules, run stats
    Sheet 2 – HHS source data in full (with ID column)
    Sheet 3 – All processed / deduplicated results + search summary
    Sheet 4 – Actionable rows: valid email addresses only
    """
    base     = os.path.splitext(scraped_csv_path)[0]
    out_path = base + ".xlsx"

    # ── load data ──────────────────────────────────────────────────────────
    hhs_rows, hhs_fields = [], []
    if os.path.exists(hhs_csv_path):
        with open(hhs_csv_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            raw_headers = next(reader, [])
            CLEAN = {0: "Name of Covered Entity", 7: "Business Associate Present"}
            hhs_fields = [CLEAN.get(i, h) for i, h in enumerate(raw_headers)]
            hhs_rows   = [dict(zip(hhs_fields, row)) for row in reader]

    with open(scraped_csv_path, newline="", encoding="utf-8-sig") as f:
        scraped_rows = list(csv.DictReader(f))

    # Deduplicate: one row per provider, keep highest-scoring
    best = {}
    for r in scraped_rows:
        key = r["provider_name"]
        if key not in best or _row_score(r) > _row_score(best[key]):
            best[key] = r
    deduped    = list(best.values())
    actionable = [r for r in deduped if _is_valid_email(r.get("emails", ""))]
    no_contact = [r for r in deduped if not _is_valid_email(r.get("emails", ""))]

    # Count how many deduped rows had an email that got junk-filtered
    # (i.e. had some email string but still failed _is_valid_email)
    junk_filtered = sum(
        1 for r in deduped
        if r.get("emails","").strip() and not _is_valid_email(r.get("emails",""))
    )

    wb = Workbook()

    # ── Sheet 1: README ───────────────────────────────────────────────────
    stats = {
        "hhs_total":     len(hhs_rows),
        "scraped_raw":   len(scraped_rows),
        "deduped":       len(deduped),
        "actionable":    len(actionable),
        "no_contact":    len(no_contact),
        "junk_filtered": junk_filtered,
        "run_time":      datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    # wb.active is sheet index 0 — we'll use it for HHS data;
    # README is inserted at position 0, pushing HHS to position 1.
    build_readme_sheet(wb, stats)

    # ── Sheet 2: HHS source ───────────────────────────────────────────────
    # wb.active is now the README sheet (index 0). The original blank default
    # sheet is at index 1 — rename and reuse it to avoid a spurious empty sheet.
    ws2 = wb.worksheets[1]
    ws2.title = "2 - HHS Source Data"
    ws2.sheet_properties.tabColor = _STEEL

    hdr2 = ["ID"] + hhs_fields
    _banner(ws2, "HHS Breach Portal — Downloaded Source Data", len(hdr2), _NAVY)
    for c, h in enumerate(hdr2, 1):
        ws2.cell(row=2, column=c).value = h
    _style_hdr(ws2, 2, len(hdr2), _STEEL)
    ws2.row_dimensions[2].height = 30

    for i, row in enumerate(hhs_rows, 1):
        rn = i + 2
        bg = _WHITE if i % 2 else _LGRAY
        ws2.cell(row=rn, column=1).value = i
        for c, field in enumerate(hhs_fields, 2):
            ws2.cell(row=rn, column=c).value = row.get(field, "")
        _style_data(ws2, rn, len(hdr2), bg)
        ws2.cell(row=rn, column=1).alignment = _ctr()

    foot2 = len(hhs_rows) + 3
    _footer(ws2, f"Total records: {len(hhs_rows)}", foot2, len(hdr2), _NAVY)
    _widths(ws2, [5] + [max(14, min(len(f) + 2, 32)) for f in hhs_fields])
    ws2.freeze_panes = "A3"

    # ── Sheet 3: all processed results ────────────────────────────────────
    ws3 = wb.create_sheet("3 - Processed Results")
    ws3.sheet_properties.tabColor = _GRN_HDR

    SCOLS = [
        ("ID", 5), ("Provider Name", 32), ("Breach Date", 13),
        ("First Name", 12), ("Last Name", 12), ("Email(s)", 34),
        ("Phone(s)", 22), ("Website", 28), ("Source URL", 36),
        ("Found Via", 13), ("Search Summary", 42), ("Date Scraped", 18),
    ]
    _banner(ws3, "Breach Providers — Processed Contact Search Results", len(SCOLS), _GRN_HDR)
    for c, (h, _) in enumerate(SCOLS, 1):
        ws3.cell(row=2, column=c).value = h
    _style_hdr(ws3, 2, len(SCOLS), _GRN_HDR)
    ws3.row_dimensions[2].height = 30

    for i, row in enumerate(deduped, 1):
        rn = i + 2
        bg = _GRN_BG if _is_valid_email(row.get("emails", "")) else _AMB_BG
        vals = [i, row.get("provider_name",""), row.get("breach_submission_date",""),
                row.get("first_name",""), row.get("last_name",""),
                row.get("emails",""), row.get("phones",""),
                row.get("website",""), row.get("source_url",""),
                row.get("found_via",""), _scrape_summary(row), row.get("date_scraped","")]
        for c, v in enumerate(vals, 1):
            ws3.cell(row=rn, column=c).value = v
        _style_data(ws3, rn, len(SCOLS), bg)
        ws3.cell(row=rn, column=1).alignment = _ctr()

    foot3 = len(deduped) + 3
    _footer(ws3,
        f"Unique providers: {len(deduped)}   |   Valid email: {len(actionable)}   |   No contact: {len(no_contact)}   |   Junk-filtered: {junk_filtered}",
        foot3, len(SCOLS), _GRN_HDR)
    _widths(ws3, [w for _, w in SCOLS])
    ws3.freeze_panes = "A3"

    # ── Sheet 4: actionable — valid email only ─────────────────────────────
    ws4 = wb.create_sheet("4 - Valid Email Contacts")
    ws4.sheet_properties.tabColor = "C00000"

    _banner(ws4, "Actionable Contacts — Valid Email Addresses Found", len(SCOLS), _RED_HDR)
    for c, (h, _) in enumerate(SCOLS, 1):
        ws4.cell(row=2, column=c).value = h
    _style_hdr(ws4, 2, len(SCOLS), _RED_HDR)
    ws4.row_dimensions[2].height = 30

    for i, row in enumerate(actionable, 1):
        rn = i + 2
        bg = _WHITE if i % 2 else _RED_BG
        vals = [i, row.get("provider_name",""), row.get("breach_submission_date",""),
                row.get("first_name",""), row.get("last_name",""),
                row.get("emails",""), row.get("phones",""),
                row.get("website",""), row.get("source_url",""),
                row.get("found_via",""), _scrape_summary(row), row.get("date_scraped","")]
        for c, v in enumerate(vals, 1):
            ws4.cell(row=rn, column=c).value = v
        _style_data(ws4, rn, len(SCOLS), bg)
        ws4.cell(row=rn, column=1).alignment = _ctr()

    foot4 = len(actionable) + 3
    _footer(ws4, f"Actionable records with valid email: {len(actionable)}", foot4, len(SCOLS), _RED_HDR)
    _widths(ws4, [w for _, w in SCOLS])
    ws4.freeze_panes = "A3"

    wb.save(out_path)
    return out_path

# -----------------------------
# MAIN PIPELINE
# -----------------------------
def main():
    print("Starting full pipeline...\n")

    init_csv()

    # Download (or reuse) the HHS CSV export and extract provider names
    providers = get_healthcare_providers()

    if not providers:
        print("❌ No providers found")
        return

    print(f"\n✅ {len(providers)} providers found:")
    for name, breach_date in providers:
        print(f" - {name} ({breach_date})")

    print("\n🔍 Finding privacy officer contacts...\n")

    for provider, breach_date in providers:
        print(f"\n{'='*60}")
        print(f"  {provider}  [{breach_date}]")
        print(f"{'='*60}")

        # Step 1: Find the official site
        site = find_official_site(provider)
        if not site:
            print("  ❌ No official site found")
            write_not_found(provider, breach_date)
            continue

        print(f"  🌐 Website: {site}")

        found_anything = False
        scrape_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Step 2: Check known privacy-related paths on the site
        contacts = find_privacy_contact(site, provider)

        for c in contacts:
            if c["emails"] or c["phones"]:
                found_anything = True
                print(f"\n  ✅ Contact found at: {c['source']}")
                if c["emails"]:
                    print(f"     📧 Emails: {', '.join(c['emails'])}")
                if c["phones"]:
                    print(f"     📞 Phones: {', '.join(c['phones'])}")
                print(f"     📝 Context: ...{c['snippet'][:200]}...")

                # Last-resort web search for name if snippet + email both failed
                first, last = c["first_name"], c["last_name"]
                if not (first and last) and c["emails"]:
                    first, last = search_for_contact_name(provider, c["emails"][0])

                write_csv_row({
                    "provider_name": provider,
                    "breach_submission_date": breach_date,
                    "first_name": first,
                    "last_name": last,
                    "website": site,
                    "emails": "; ".join(c["emails"]),
                    "phones": "; ".join(c["phones"]),
                    "source_url": c["source"],
                    "context_snippet": c["snippet"][:300],
                    "found_via": "site crawl",
                    "date_scraped": scrape_time,
                })

        # Step 3: If nothing found on-site, try a targeted web search
        if not found_anything:
            print(f"\n  🔎 Trying targeted web search for privacy officer...")
            emails, source, first, last = search_for_privacy_officer(provider)
            if emails:
                print(f"  ✅ Found via web search!")
                print(f"     📧 Emails: {', '.join(emails)}")
                if first or last:
                    print(f"     👤 Name: {first} {last}".strip())
                print(f"     Source: {source}")

                write_csv_row({
                    "provider_name": provider,
                    "breach_submission_date": breach_date,
                    "first_name": first,
                    "last_name": last,
                    "website": site,
                    "emails": "; ".join(emails),
                    "phones": "",
                    "source_url": source,
                    "context_snippet": "",
                    "found_via": "web search",
                    "date_scraped": scrape_time,
                })
            else:
                print(f"  ❌ No privacy officer contact found")
                write_not_found(provider, breach_date, site)

    print(f"\n✅ Done. Results saved to: {CSV_FILE}")

    # Build Excel report automatically
    print("\n📊 Building Excel report...")
    try:
        xlsx_path = build_excel_report(HHS_CSV_PATH, CSV_FILE)
        print(f"📗 Excel report saved to: {xlsx_path}")
    except Exception as e:
        print(f"⚠️  Excel report failed: {e}")


if __name__ == "__main__":
    main()
