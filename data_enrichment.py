import pandas as pd
import re
import argparse
from openpyxl import load_workbook

DEFAULT_VALUE = "Undetermined"


class DataEnricher:

	FREE_EMAIL_DOMAINS = {
		'gmail.com','yahoo.com','outlook.com','hotmail.com',
		'aol.com','protonmail.com'
	}

	DOMAIN_ORG_TYPES = {
		'gov':'Government',
		'edu':'Education',
		'mil':'Military',
		'org':'Non-Profit'
	}

	COUNTRY_TLDS = {
		'uk':'United Kingdom','de':'Germany','fr':'France','it':'Italy',
		'es':'Spain','nl':'Netherlands','se':'Sweden','ch':'Switzerland',
		'au':'Australia','ca':'Canada','jp':'Japan','cn':'China',
		'in':'India','br':'Brazil','mx':'Mexico','us':'United States',
		'ru':'Russia','za':'South Africa'
	}

	US_STATES = {
		'AL','AK','AZ','AR','CA','CO','CT','DE','FL','GA','HI','ID',
		'IL','IN','IA','KS','KY','LA','ME','MD','MA','MI','MN','MS',
		'MO','MT','NE','NV','NH','NJ','NM','NY','NC','ND','OH','OK',
		'OR','PA','RI','SC','SD','TN','TX','UT','VT','VA','WA','WV',
		'WI','WY','DC'
	}


	# -------------------------
	# Utilities
	# -------------------------

	@staticmethod
	def clean_value(value):

		if pd.isna(value):
			return DEFAULT_VALUE

		v = str(value).strip()

		if v == "":
			return DEFAULT_VALUE

		return v


	@staticmethod
	def extract_domain(email):

		email = DataEnricher.clean_value(email)

		if email == DEFAULT_VALUE or '@' not in email:
			return DEFAULT_VALUE

		return email.split('@')[-1].lower()


	@staticmethod
	def extract_area_code(phone):

		phone = DataEnricher.clean_value(phone)

		if phone == DEFAULT_VALUE:
			return DEFAULT_VALUE

		digits = re.sub(r'\D','',phone)

		if len(digits) == 10:
			return digits[:3]

		if digits.startswith('1') and len(digits) == 11:
			return digits[1:4]

		return DEFAULT_VALUE


	def extract_primary_domain(self,row):

		emails = [
			row.get('Person - Email - Work'),
			row.get('Person - Email - Home'),
			row.get('Person - Email - Other')
		]

		for email in emails:

			domain = self.extract_domain(email)

			if domain != DEFAULT_VALUE:
				return domain

		return DEFAULT_VALUE


	# -------------------------
	# US Detection
	# -------------------------

	def check_state_us(self,row):

		state = self.clean_value(row.get('Person - State'))

		if state == DEFAULT_VALUE:
			return None

		state = state.upper()

		if state in self.US_STATES:
			return "US"

		return None


	def check_email_us(self,row):

		emails = [
			row.get('Person - Email - Work'),
			row.get('Person - Email - Home'),
			row.get('Person - Email - Other')
		]

		for email in emails:

			domain = self.extract_domain(email)

			if domain == DEFAULT_VALUE:
				continue

			tld = domain.split('.')[-1]

			if tld == "us":
				return "US"

			if tld in self.COUNTRY_TLDS and tld != "us":
				return "Non-US"

		return None


	def check_phone_us(self,row):

		phones = [
			row.get('Person - Phone - Work'),
			row.get('Person - Phone - Home')
		]

		for phone in phones:

			area = self.extract_area_code(phone)

			if area == DEFAULT_VALUE:
				continue

			try:
				num = int(area)

				if 200 <= num < 1000:
					return "US"

			except:
				pass

		return None


	def determine_us_status(self,row):

		state_check = self.check_state_us(row)

		if state_check:
			return state_check

		email_check = self.check_email_us(row)

		if email_check:
			return email_check

		phone_check = self.check_phone_us(row)

		if phone_check:
			return phone_check

		return "US Likely"  # Default to US Likely if no clear indicators


	# -------------------------
	# Enrichment Logic
	# -------------------------

	def determine_country(self,domain):

		if domain == DEFAULT_VALUE:
			return DEFAULT_VALUE

		ext = domain.split('.')[-1]

		if ext in self.COUNTRY_TLDS:
			return self.COUNTRY_TLDS[ext]

		return DEFAULT_VALUE


	def determine_location(self,row):

		state = self.clean_value(row.get('Person - State'))

		if state != DEFAULT_VALUE:

			state = state.upper()

			if state in self.US_STATES:
				return state

		return DEFAULT_VALUE


	def determine_org_type(self,row,domain):

		if domain == DEFAULT_VALUE:
			return DEFAULT_VALUE

		if domain in self.FREE_EMAIL_DOMAINS:
			return "Personal"

		tld = domain.split('.')[-1]

		if tld in self.DOMAIN_ORG_TYPES:
			return self.DOMAIN_ORG_TYPES[tld]

		org_name = self.clean_value(row.get('Person - CompanyName'))

		if org_name != DEFAULT_VALUE:

			org = org_name.lower()

			if any(x in org for x in ['university','college','school']):
				return "Education"

			if any(x in org for x in ['hospital','clinic','medical']):
				return "Healthcare"

			if any(x in org for x in ['foundation','charity','nonprofit']):
				return "Non-Profit"

			if any(x in org for x in ['inc','corp','llc','ltd']):
				return "Corporate"

		return "Other"


	def infer_company_domain(self,domain):

		if domain == DEFAULT_VALUE:
			return DEFAULT_VALUE

		if domain in self.FREE_EMAIL_DOMAINS:
			return DEFAULT_VALUE

		return domain


	def infer_website(self,row,domain):

		website = self.clean_value(row.get('Person - Website'))

		if website != DEFAULT_VALUE:

			if not website.startswith(('http://','https://')):
				return "https://" + website

			return website

		if domain != DEFAULT_VALUE and domain not in self.FREE_EMAIL_DOMAINS:
			return f"https://www.{domain}"

		return DEFAULT_VALUE


	def determine_lead_type(self,domain):

		if domain == DEFAULT_VALUE:
			return DEFAULT_VALUE

		if domain in self.FREE_EMAIL_DOMAINS:
			return "Personal"

		return "Business"


	def score_lead(self,row):

		score = 0

		if not row['Free Email']:
			score += 3

		if row['Organization Type'] == "Corporate":
			score += 3

		if row['Website'] != DEFAULT_VALUE:
			score += 2

		if row['Phone Area Code'] != DEFAULT_VALUE:
			score += 1

		if row['Country'] != DEFAULT_VALUE:
			score += 1

		return score


	# -------------------------
	# Pipeline
	# -------------------------

	def add_enrichment_columns(self,df):

		print("Running enrichment pipeline...")

		df = df.fillna(DEFAULT_VALUE)

		df['Email Domain'] = df.apply(self.extract_primary_domain, axis=1)

		df['Free Email'] = df['Email Domain'].isin(self.FREE_EMAIL_DOMAINS)

		df['Phone Area Code'] = df['Person - Phone - Work'].apply(self.extract_area_code)

		df['Company Domain'] = df['Email Domain'].apply(self.infer_company_domain)

		df['Country'] = df['Email Domain'].apply(self.determine_country)

		df['US Status'] = df.apply(self.determine_us_status, axis=1)

		df['Location'] = df.apply(self.determine_location,axis=1)

		df['Organization Type'] = df.apply(
			lambda r: self.determine_org_type(r,r['Email Domain']),
			axis=1
		)

		df['Website'] = df.apply(
			lambda r: self.infer_website(r,r['Email Domain']),
			axis=1
		)

		df['Lead Type'] = df['Email Domain'].apply(self.determine_lead_type)

		df['Lead Score'] = df.apply(self.score_lead,axis=1)

		return df


# -------------------------
# File Processing
# -------------------------

def process_csv(input_csv,output_csv):

	df = pd.read_csv(input_csv)

	enricher = DataEnricher()

	df = enricher.add_enrichment_columns(df)

	df.to_csv(output_csv,index=False)

	print("CSV created")


def process_xlsx(input_csv,output_xlsx):

	df = pd.read_csv(input_csv)

	enricher = DataEnricher()

	df = enricher.add_enrichment_columns(df)

	df.to_excel(output_xlsx,index=False)

	print("Excel created")


def update_excel(excel_file,input_csv):

	df = pd.read_csv(input_csv)

	enricher = DataEnricher()

	df = enricher.add_enrichment_columns(df)

	wb = load_workbook(excel_file)

	ws = wb.active

	start_row = 2

	if ws.max_row >= start_row:
		ws.delete_rows(start_row, ws.max_row - start_row + 1)

	for c,col in enumerate(df.columns,start=1):
		ws.cell(row=1,column=c,value=col)

	for r,row in enumerate(df.values,start=start_row):
		for c,value in enumerate(row,start=1):

			if pd.isna(value) or str(value).strip() == "":
				value = DEFAULT_VALUE

			ws.cell(row=r,column=c,value=value)

	wb.save(excel_file)

	print("Excel updated")


# -------------------------
# CLI
# -------------------------

if __name__ == "__main__":

	parser = argparse.ArgumentParser(description="CRM Contact Enrichment Tool")

	sub = parser.add_subparsers(dest="command")

	csv_cmd = sub.add_parser("csv")
	csv_cmd.add_argument("input_csv")
	csv_cmd.add_argument("output_csv")

	xlsx_cmd = sub.add_parser("xlsx")
	xlsx_cmd.add_argument("input_csv")
	xlsx_cmd.add_argument("output_xlsx")

	update_cmd = sub.add_parser("update")
	update_cmd.add_argument("input_csv")
	update_cmd.add_argument("excel_file")

	args = parser.parse_args()

	if args.command == "csv":
		process_csv(args.input_csv,args.output_csv)

	elif args.command == "xlsx":
		process_xlsx(args.input_csv,args.output_xlsx)

	elif args.command == "update":
		update_excel(args.excel_file,args.input_csv)

	else:
		parser.print_help()